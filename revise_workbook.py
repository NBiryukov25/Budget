#!/usr/bin/env python3
"""Revise an existing survival workbook in place. Never rebuilds it.

Loads the user's live file, edits it, and writes a new version. The cash-flow
model, financial data, recurring logic, visual style and working formulas are
preserved; only the things the change list calls for are touched.

    python3 revise_workbook.py <in.xlsx> <out.xlsx>
"""

import argparse
import datetime as dt
import re

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ------------------------------------------------------------------ geometry
SLOTS, FIRST_BLOCK, N_WEEKS = 10, 13, 13
REC_FIRST, REC_LAST, REC_HDR = 5, 34, 4
SCH_FIRST = 4
LOG_HDR, LOG_FIRST, LOG_LAST = 8, 9, 308
OVR_HDR, OVR_FIRST, OVR_LAST = 8, 9, 108

CUR, PLOG, OVR, SCH = "'Cash Flow'", "'Paid Log'", "'Overrides'", "Schedule"

# -------------------------------------------------------------------- styles
FONT = "Arial"
F_TITLE = Font(name=FONT, size=16, bold=True, color="1F3864")
F_SUB = Font(name=FONT, size=10, italic=True, color="595959")
F_H1 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_H2 = Font(name=FONT, size=10, bold=True, color="1F3864")
F_BODY = Font(name=FONT, size=10, color="000000")
F_INPUT = Font(name=FONT, size=10, color="0000FF")
F_LINK = Font(name=FONT, size=10, color="008000")
F_BOLD = Font(name=FONT, size=10, bold=True, color="000000")
F_NOTE = Font(name=FONT, size=9, italic=True, color="595959")
F_BIG = Font(name=FONT, size=14, bold=True, color="1F3864")

FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_SUB = PatternFill("solid", fgColor="D9E2F3")
FILL_IN = PatternFill("solid", fgColor="FFFF00")
FILL_TOT = PatternFill("solid", fgColor="E2EFDA")
FILL_BAND = PatternFill("solid", fgColor="F2F2F2")
FILL_CARD = PatternFill("solid", fgColor="F7F9FC")

# High-contrast status chips. Dark ink on a pale ground of the same hue - the
# pairs Excel itself ships for Good / Neutral / Bad, so nothing can end up
# reading as text the same colour as its background.
CHIP = {
    "good":    ("006100", "C6EFCE"),
    "warn":    ("9C5700", "FFEB9C"),
    "bad":     ("9C0006", "FFC7CE"),
    "info":    ("1F4E78", "DDEBF7"),
    "neutral": ("3F3F76", "EDEDED"),
}

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '$#,##0.00;($#,##0.00);"–"'
MONEY_B = '$#,##0.00;[Red]($#,##0.00);"–"'
DATE_F = "ddd mm/dd"
DATE_L = "ddd mmm d, yyyy"

UNLOCKED = Protection(locked=False)
LOCKED = Protection(locked=True)


def put(ws, ref, value, font=F_BODY, fmt=None, fill=None, align=None,
        border=None, wrap=False, unlocked=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = border
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.protection = UNLOCKED if unlocked else LOCKED
    return c


def chip_rules(ws, rng, mapping):
    """One conditional rule per status word, dark ink on a pale ground."""
    for word, kind in mapping.items():
        ink, ground = CHIP[kind]
        first = rng.split(":")[0]
        col = re.match(r"\$?([A-Z]+)", first).group(1)
        row = re.search(r"(\d+)", first).group(1)
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f'${col}{row}="{word}"'],
                             font=Font(name=FONT, size=10, bold=True, color=ink),
                             fill=PatternFill("solid", fgColor=ground)))


def detect_var_rows(cf):
    for vr in range(4, 41):
        h2 = FIRST_BLOCK + (SLOTS + vr + 5)
        v = cf[f"B{h2}"].value
        if isinstance(v, str) and v.startswith('="WEEK '):
            return vr
    raise SystemExit("cannot determine block height")


def detect_sched_last(sch):
    r = SCH_FIRST
    while sch[f"A{r}"].value is not None:
        r += 1
    return r - 1


def block(w, vr):
    h = FIRST_BLOCK + (SLOTS + vr + 5) * (w - 1)
    return {"head": h, "cols": h + 1, "fs": h + 2, "ls": h + 1 + SLOTS,
            "vhdr": h + 2 + SLOTS, "fv": h + 3 + SLOTS,
            "lv": h + 2 + SLOTS + vr, "tot": h + 3 + SLOTS + vr}


def clear_sheet(ws, last_row, last_col):
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))
    for row in ws.iter_rows(min_row=1, max_row=last_row, max_col=last_col):
        for c in row:
            c.value = None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src"); ap.add_argument("out")
    a = ap.parse_args()

    wb = load_workbook(a.src)
    vals = load_workbook(a.src, data_only=True)
    cf, rec, sm, plog, sch = (wb["Cash Flow"], wb["Recurring"], wb["Summary"],
                              wb["Paid Log"], wb["Schedule"])
    cfv = vals["Cash Flow"]
    VR = detect_var_rows(cf)
    SCH_LAST = detect_sched_last(sch)
    notes = []

    LOG_D = f"{PLOG}!$B${LOG_FIRST}:$B${LOG_LAST}"
    LOG_N = f"{PLOG}!$C${LOG_FIRST}:$C${LOG_LAST}"
    LOG_A = f"{PLOG}!$D${LOG_FIRST}:$D${LOG_LAST}"
    O_N = f"{OVR}!$B${OVR_FIRST}:$B${OVR_LAST}"
    O_OD = f"{OVR}!$C${OVR_FIRST}:$C${OVR_LAST}"
    O_ND = f"{OVR}!$D${OVR_FIRST}:$D${OVR_LAST}"
    O_AM = f"{OVR}!$E${OVR_FIRST}:$E${OVR_LAST}"

    # ============================================ 16 · named assumption cells
    for name, ref in (("BeginningBalance", "$E$5"), ("StartDate", "$E$6"),
                      ("MinCushion", "$E$7"), ("WeeklySpend", "$E$8")):
        if name in wb.defined_names:
            del wb.defined_names[name]
        from openpyxl.workbook.defined_name import DefinedName
        wb.defined_names.add(DefinedName(name, attr_text=f"'Cash Flow'!{ref}"))

    # ==================================== 1 · start date: describe, never scold
    put(cf, "G6",
        '=IF($E$6="","","Projection runs "&TEXT($E$6,"ddd mmm d")&" to "'
        f'&TEXT($E$6+{N_WEEKS * 7 - 1},"ddd mmm d, yyyy")&"  ·  periods are '
        'full 7-day weeks beginning "&TEXT($E$6,"dddd")&"s.")',
        F_NOTE)
    notes.append("Start date: the Friday-only warning is gone; any start day is accepted "
                 "and the note now just states the window and which weekday periods begin on.")

    # ================================================ 5 · Overrides worksheet
    if "Overrides" in wb.sheetnames:
        del wb["Overrides"]
    ovr = wb.create_sheet("Overrides", wb.sheetnames.index("Paid Log") + 1)
    ovr.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 34, "C": 17, "D": 17, "E": 15, "F": 40}.items():
        ovr.column_dimensions[col].width = w
    put(ovr, "B1", "OVERRIDES", F_TITLE)
    put(ovr, "B2", "Change one occurrence without touching the recurring schedule.", F_SUB)
    for i, line in enumerate([
        "Name the transaction exactly as it reads on the Cash Flow tab and give the date it "
        "is currently scheduled for. Fill in a new date, a new amount, or both.",
        "Leave a field blank and that part is not overridden. Delete the row and the "
        "transaction goes straight back to its normal schedule — nothing else to undo.",
        "This is for a one-off change. To move a bill permanently, edit its Next Due Date "
        "on the Recurring tab instead.",
    ]):
        put(ovr, f"B{4 + i}", line, F_NOTE, wrap=True)
        ovr.merge_cells(f"B{4+i}:F{4+i}")
        ovr.row_dimensions[4 + i].height = 24
    for col, label in (("B", "Transaction"), ("C", "Scheduled Date"),
                       ("D", "Use This Date Instead"), ("E", "Use This Amount"),
                       ("F", "Why")):
        put(ovr, f"{col}{OVR_HDR}", label, F_H1, fill=FILL_HDR, align="center", border=BOX)
    for r in range(OVR_FIRST, OVR_LAST + 1):
        band = FILL_BAND if (r - OVR_HDR) % 2 == 0 else None
        for col in "BCDEF":
            put(ovr, f"{col}{r}", None, F_INPUT, fill=band, border=BOX, unlocked=True)
        for col in "CD":
            ovr[f"{col}{r}"].number_format = DATE_L
            ovr[f"{col}{r}"].alignment = Alignment(horizontal="center")
        ovr[f"E{r}"].number_format = MONEY
    ovr.freeze_panes = f"B{OVR_FIRST}"
    notes.append("Overrides: new sheet. Give a transaction a different date or amount for one "
                 "occurrence; clear the row to revert. No formula is ever overwritten.")

    # ===================== 5 + 18 · schedule engine honours date/amount overrides
    for r in range(SCH_FIRST, SCH_LAST + 1):
        g_old = sch[f"G{r}"].value
        if not isinstance(g_old, str):
            continue
        # The weekend shift moves to L; G becomes the override-aware due date so
        # every downstream week number, ranking and Cash Flow pull follows it.
        sch[f"L{r}"] = g_old
        sch[f"L{r}"].number_format = DATE_L
        sch[f"L{r}"].font = F_LINK
        newdate = f'SUMIFS({O_ND},{O_N},$C{r},{O_OD},$L{r})'
        sch[f"G{r}"] = (f'=IF($L{r}="","",IF({newdate}>0,{newdate},$L{r}))')
        e_old = sch[f"E{r}"].value
        m = re.search(r"Recurring!\$D\$(\d+)", str(e_old) or "")
        if m:
            amt_ovr = f'SUMIFS({O_AM},{O_N},$C{r},{O_OD},$L{r})'
            sch[f"E{r}"] = (f'=IF($C{r}="","",IF({amt_ovr}>0,{amt_ovr},'
                            f'Recurring!$D${m.group(1)}))')
        # helper keys: next big bill, next money in (unique so MATCH is exact)
        sch[f"M{r}"] = (f'=IF(AND($H{r}=1,$D{r}="Expense",$G{r}>TODAY(),$E{r}>=100),'
                        f'$G{r}*100000+ROW(),"")')
        sch[f"N{r}"] = (f'=IF(AND($H{r}=1,$D{r}="Income",$G{r}>TODAY()),'
                        f'$G{r}*100000+ROW(),"")')
    put(sch, "L3", "Base Date", F_H1, fill=FILL_HDR, align="center", border=BOX)
    put(sch, "M3", "Big Bill Key", F_H1, fill=FILL_HDR, align="center", border=BOX)
    put(sch, "N3", "Income Key", F_H1, fill=FILL_HDR, align="center", border=BOX)
    for col, w in {"L": 15, "M": 14, "N": 14}.items():
        sch.column_dimensions[col].width = w
    SCH_M = f"{SCH}!$M${SCH_FIRST}:$M${SCH_LAST}"
    SCH_N = f"{SCH}!$N${SCH_FIRST}:$N${SCH_LAST}"
    SCH_C = f"{SCH}!$C${SCH_FIRST}:$C${SCH_LAST}"
    SCH_E = f"{SCH}!$E${SCH_FIRST}:$E${SCH_LAST}"

    # =================== 6 + 13 · recurring validation, filtering, input styling
    put(rec, f"K{REC_HDR}", "Needs Attention", F_H1, fill=FILL_HDR,
        align="center", border=BOX, wrap=True)
    for r in range(REC_FIRST, REC_LAST + 1):
        rec[f"K{r}"] = (
            f'=IF($B{r}="","",IF($H{r}<>"Yes","",'
            f'IF(OR($D{r}="",$D{r}=0),"! No amount — this will project as $0",'
            f'IF($F{r}="","! No next due date",'
            f'IF(NOT(ISNUMBER($F{r})),"! Next due date is not a date — type it as 8/21/2026",'
            f'IF($E{r}="","! No frequency chosen",'
            f'IF($F{r}<StartDate-90,"! "&TEXT($F{r},"mmm d, yyyy")&" is too far back to appear",'
            f'IF($F{r}>StartDate+730,"! "&TEXT($F{r},"mmm d, yyyy")&" is years away — check it",'
            f'IF($F{r}>StartDate+{N_WEEKS * 7 - 1},"beyond the {N_WEEKS}-week window","")))))))) ')
        rec[f"K{r}"].font = Font(name=FONT, size=9, bold=True, color="9C0006")
        rec[f"K{r}"].border = BOX
        for col in "BCDEFGHIJ":
            rec[f"{col}{r}"].protection = UNLOCKED
    rec.auto_filter.ref = f"A{REC_HDR}:K{REC_LAST}"
    rec.column_dimensions["K"].width = 46
    notes.append("Recurring: the check column now also catches an active item with no amount, "
                 "no frequency or no date, and the header row filters and sorts.")

    # ============================================= 4 · Paid Log key made visible
    put(plog, f"F{LOG_HDR}", "Match Key", F_H1, fill=FILL_HDR, align="center", border=BOX)
    plog.column_dimensions["F"].width = 40
    for r in range(LOG_FIRST, LOG_LAST + 1):
        plog[f"F{r}"] = (f'=IF($C{r}="","",$C{r}&"  |  "&TEXT($B{r},"yyyy-mm-dd"))')
        plog[f"F{r}"].font = F_NOTE
        plog[f"F{r}"].border = BOX
        for col in "BCDE":
            plog[f"{col}{r}"].protection = UNLOCKED
    put(plog, f"B{LOG_HDR - 1}",
        "This pairing of description and scheduled date is what the Cash Flow looks up, so a "
        "logged payment stays with its own transaction however the schedule is re-sorted.",
        F_NOTE, wrap=True)
    plog.merge_cells(f"B{LOG_HDR-1}:F{LOG_HDR-1}")
    plog.row_dimensions[LOG_HDR - 1].height = 24

    print(f"geometry: {VR} variable rows/week, schedule rows {SCH_FIRST}-{SCH_LAST}")
    wb.save(a.out)
    print("stage 1 saved ->", a.out)
    for n in notes:
        print("  *", n)


if __name__ == "__main__":
    main()
