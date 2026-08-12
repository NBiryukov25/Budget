#!/usr/bin/env python3
"""Carry a working copy's entered data into a freshly generated workbook.

The week blocks are a fixed geometry, so changing how many rows a section holds
moves every row below it. Inserting rows in place would leave every formula
pointing at the wrong cells; instead the workbook is regenerated at the new size
and everything typed by hand is copied across by position.

What moves: the three input cells, the whole Recurring table, each week's
Amount Paid / Received entries (matched by slot number), and each week's
variable-expense rows (description, date, amount). Formulas are never copied -
the new workbook brings its own.

    python3 migrate_workbook.py <old.xlsx> <new_blank.xlsx> <out.xlsx> \
        --old-var-rows 6 --new-var-rows 10
"""

import argparse

from openpyxl import load_workbook

N_WEEKS, SLOTS, FIRST_BLOCK = 13, 10, 13
REC_FIRST, REC_LAST = 5, 34


def geometry(var_rows):
    block_h = SLOTS + var_rows + 5

    def rows(w):
        h = FIRST_BLOCK + block_h * (w - 1)
        return {"head": h, "first_slot": h + 2,
                "first_var": h + 3 + SLOTS, "last_var": h + 2 + SLOTS + var_rows}

    return rows


def is_typed(cell):
    return cell.value is not None and not str(cell.value).startswith("=")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("old")
    ap.add_argument("blank")
    ap.add_argument("out")
    ap.add_argument("--old-var-rows", type=int, required=True)
    ap.add_argument("--new-var-rows", type=int, required=True)
    a = ap.parse_args()

    old = load_workbook(a.old)                       # formulas intact
    oldv = load_workbook(a.old, data_only=True)      # for date values
    new = load_workbook(a.blank)

    o_geo, n_geo = geometry(a.old_var_rows), geometry(a.new_var_rows)
    ocf, ocfv, ncf = old["Cash Flow"], oldv["Cash Flow"], new["Cash Flow"]

    # input cells - E5 may itself be a formula the user typed, so copy verbatim
    moved = []
    for ref in ("E5", "E6", "E7"):
        ncf[ref] = ocf[ref].value
        moved.append(f"{ref}={ocf[ref].value}")

    # the whole recurring table, in case rows were edited or added
    orec, nrec = old["Recurring"], new["Recurring"]
    rec_rows = 0
    for r in range(REC_FIRST, REC_LAST + 1):
        if orec[f"B{r}"].value is None:
            continue
        rec_rows += 1
        for col in "BCDEFGHIJ":
            nrec[f"{col}{r}"] = orec[f"{col}{r}"].value

    paid_n = var_n = 0
    for w in range(1, N_WEEKS + 1):
        o, n = o_geo(w), n_geo(w)

        for i in range(SLOTS):                        # Amount Paid / Received
            src = ocf[f"F{o['first_slot'] + i}"]
            if is_typed(src):
                ncf[f"F{n['first_slot'] + i}"] = src.value
                paid_n += 1

        for i in range(a.old_var_rows):               # variable expense rows
            r_o = o["first_var"] + i
            if not is_typed(ocf[f"B{r_o}"]):
                continue
            if i >= a.new_var_rows:
                raise SystemExit(f"week {w} row {i} would not fit the new block")
            r_n = n["first_var"] + i
            ncf[f"B{r_n}"] = ocf[f"B{r_o}"].value
            ncf[f"E{r_n}"] = ocf[f"E{r_o}"].value
            date_val = ocfv[f"C{r_o}"].value
            if date_val is not None:
                ncf[f"C{r_n}"] = date_val
                ncf[f"C{r_n}"].number_format = "ddd mm/dd"
            var_n += 1

    new.save(a.out)
    print(f"inputs: {', '.join(moved)}")
    print(f"recurring rows: {rec_rows} | amount-paid entries: {paid_n} | "
          f"variable rows: {var_n}")
    print(f"-> {a.out}")


if __name__ == "__main__":
    main()
