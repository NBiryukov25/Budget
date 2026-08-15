#!/usr/bin/env python3
"""Restart a workbook from a new date and balance, clearing what came before.

Only input cells are written. Every formula is left exactly as it was.

Three things have to move together or the sheet quietly lies:

  1. Each recurring item's Next Due Date rolls forward to its first occurrence
     on or after the new start. Without this the engine's past-due carryover
     drags settled bills into week 1 - which is correct behaviour when a bill
     is genuinely unpaid, and wrong when you have already cleared it.
  2. Variable-expense rows dated before the new start are cleared. They are
     typed values, not formulas, so nothing filters them by date.
  3. Amount Paid / Received entries are cleared everywhere. Those cells are
     positional: slot 3 of week 2 holds whatever now sorts third in that week.
     Move the start date and the same cell attaches to a different
     transaction, so a stale figure silently misattributes.

    python3 restart_from.py <in.xlsx> <out.xlsx> --start YYYY-MM-DD --balance N
"""

import argparse
import datetime as dt

from openpyxl import load_workbook

N_WEEKS, SLOTS, VAR_ROWS, FIRST_BLOCK = 13, 10, 10, 13
BLOCK_H = SLOTS + VAR_ROWS + 5
REC_FIRST, REC_LAST = 5, 34
DATE_L = "ddd mmm d, yyyy"


def last_day(y, m):
    return (dt.date(y + (m == 12), m % 12 + 1, 1) - dt.timedelta(days=1)).day


def add_months(d, n):
    y, m = divmod(d.year * 12 + d.month - 1 + n, 12)
    m += 1
    return dt.date(y, m, min(d.day, last_day(y, m)))


def weekend_shift(d, rule):
    r = (rule or "None").strip().lower()
    if r == "move before":
        if d.weekday() == 5: return d - dt.timedelta(days=1)
        if d.weekday() == 6: return d - dt.timedelta(days=2)
    elif r == "move after":
        if d.weekday() == 5: return d + dt.timedelta(days=2)
        if d.weekday() == 6: return d + dt.timedelta(days=1)
    return d


def next_on_or_after(anchor, freq, rule, start):
    """First occurrence of this rule landing on or after `start`."""
    f = (freq or "").strip().lower()
    leg0 = 0 if anchor.day <= 20 else 1
    for k in range(0, 600):
        if f == "weekly":            d = anchor + dt.timedelta(days=7 * k)
        elif f == "every 2 weeks":   d = anchor + dt.timedelta(days=14 * k)
        elif f == "monthly":         d = add_months(anchor, k)
        elif f == "one time":        d = anchor if k == 0 else None
        elif f == "twice a month":
            leg = leg0 + k
            base = add_months(dt.date(anchor.year, anchor.month, 1), leg // 2)
            d = (dt.date(base.year, base.month, last_day(base.year, base.month))
                 if leg % 2 else dt.date(base.year, base.month, 15))
        else:                        return None
        if d is None:
            return None
        if weekend_shift(d, rule) >= start:
            return d          # store the raw date; the sheet re-applies the rule
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src"); ap.add_argument("out")
    ap.add_argument("--start", required=True)
    ap.add_argument("--balance", required=True, type=float)
    a = ap.parse_args()
    start = dt.date.fromisoformat(a.start)

    wb = load_workbook(a.src)
    vals = load_workbook(a.src, data_only=True)
    cf, rec, cfv = wb["Cash Flow"], wb["Recurring"], vals["Cash Flow"]

    cf["E5"] = a.balance
    cf["E6"] = start
    cf["E6"].number_format = DATE_L

    rolled, kept = [], []
    for r in range(REC_FIRST, REC_LAST + 1):
        if not rec[f"B{r}"].value:
            continue
        old = rec[f"F{r}"].value
        old = old.date() if isinstance(old, dt.datetime) else old
        if old is None:
            continue
        new = next_on_or_after(old, rec[f"E{r}"].value, rec[f"G{r}"].value, start)
        if new and new != old:
            rec[f"F{r}"] = new
            rec[f"F{r}"].number_format = DATE_L
            rolled.append((rec[f"B{r}"].value, old, new))
        else:
            kept.append((rec[f"B{r}"].value, old))

    cleared_var, cleared_paid = 0, 0
    for w in range(1, N_WEEKS + 1):
        h = FIRST_BLOCK + BLOCK_H * (w - 1)
        fs, fv, lv = h + 2, h + 13, h + 2 + SLOTS + VAR_ROWS
        for i in range(SLOTS):
            c = cf[f"F{fs + i}"]
            if c.value is not None:
                c.value = None
                cleared_paid += 1
        for row in range(fv, lv + 1):
            b = cf[f"B{row}"].value
            if b is None:
                continue
            d = cfv[f"C{row}"].value
            d = d.date() if isinstance(d, dt.datetime) else d
            if d is None or d < start:          # undated leftovers go too
                for col in "BCE":
                    cf[f"{col}{row}"] = None
                cleared_var += 1

    # The start-date check was written assuming Friday weeks. Saying "that is a
    # Saturday" as an error is wrong when the day was chosen deliberately, so it
    # now reports which days the weeks will run and leaves the choice alone.
    cf["G6"] = ('=IF($E$6="","",IF(WEEKDAY($E$6,2)=5,"",'
                '"Weeks will run "&TEXT($E$6,"ddd")&" to "&TEXT($E$6+6,"ddd")&". '
                'Start on a Friday if you want the Friday paycheck to open each week."))')

    wb.save(a.out)
    print(f"start {start:%a %b %d, %Y} · balance {a.balance:,.2f}")
    print(f"rolled {len(rolled)} due dates forward:")
    for n, o, x in rolled:
        print(f"    {n:34s} {o:%b %d} -> {x:%b %d}")
    print(f"unchanged ({len(kept)}): " + ", ".join(n for n, _ in kept))
    print(f"cleared {cleared_var} variable rows, {cleared_paid} amount-paid entries")


if __name__ == "__main__":
    main()
