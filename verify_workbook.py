#!/usr/bin/env python3
"""Item 20: prove the finished workbook behaves, rather than asserting it."""
import datetime as dt, shutil, subprocess, sys
from openpyxl import load_workbook

RECALC = "/root/.claude/skills/synced/xlsx/scripts/recalc.py"
ERRS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "Err:")
n = lambda v: v if isinstance(v, (int, float)) else 0
SLOTS, FB, NW = 10, 13, 13

def _detect(path):
    """Read the block height off the workbook instead of assuming it."""
    cf = load_workbook(path)["Cash Flow"]
    for vr in range(4, 61):
        v = cf[f"B{FB + (SLOTS + vr + 5)}"].value
        if isinstance(v, str) and v.startswith('="WEEK '):
            return vr
    raise SystemExit("cannot determine block height")

VR = _detect(sys.argv[1])
blk = lambda w: {"h": FB+(SLOTS+VR+5)*(w-1), "fs": FB+(SLOTS+VR+5)*(w-1)+2,
                 "lv": FB+(SLOTS+VR+5)*(w-1)+2+SLOTS+VR,
                 "tot": FB+(SLOTS+VR+5)*(w-1)+3+SLOTS+VR}

def recalc(p):
    subprocess.run([sys.executable, RECALC, p, "300"], capture_output=True)

def scan_errors(p):
    f, v = load_workbook(p), load_workbook(p, data_only=True)
    bad = []
    for name in f.sheetnames:
        sf, sv = f[name], v[name]
        for row in sf.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    r = sv[c.coordinate].value
                    if isinstance(r, str) and any(r.strip().startswith(e) for e in ERRS):
                        bad.append(f"{name}!{c.coordinate}={r}")
    return bad

def result(label, ok, detail=""):
    print(f"  [{'PASS' if ok else 'FAIL'}] {label}{('  — ' + detail) if detail else ''}")
    return ok

src = sys.argv[1]
allok = True
print(f"VALIDATION  ({VR} variable rows per week)")

# --- no calculation errors anywhere
bad = scan_errors(src)
allok &= result("no #REF!/#VALUE!/#NAME? anywhere", not bad, "; ".join(bad[:5]))

v = load_workbook(src, data_only=True)
cf, sm = v["Cash Flow"], v["Summary"]

# --- start date preserved exactly, non-Friday
sd = cf["E6"].value
allok &= result("start date preserved as entered", sd == dt.datetime(2026, 8, 15),
                f"{sd:%a %b %d, %Y}")

# --- Summary agrees with Cash Flow, week by week
mism = [w for w in range(1, NW+1)
        if abs(n(sm[f"I{12+w}"].value) - n(cf[f"I{blk(w)['tot']}"].value)) > 0.005
        or abs(n(sm[f"H{12+w}"].value) - n(cf[f"I{blk(w)['h']}"].value)) > 0.005]
allok &= result("Summary ending/carried match Cash Flow, all 13 weeks", not mism, str(mism))

# --- dashboard agrees with the table
allok &= result("dashboard lowest == MIN of weekly lows",
                abs(n(sm["D6"].value) - min(n(sm[f"J{12+w}"].value) for w in range(1, NW+1))) < 0.005)
allok &= result("dashboard ending == week 13 ending",
                abs(n(sm["D7"].value) - n(sm[f"I{25}"].value)) < 0.005)

# --- status text is a real word, never blank-on-blank
words = {cf[f"J{r}"].value for w in range(1, NW+1)
         for r in range(blk(w)["fs"], blk(w)["tot"]+1) if cf[f"J{r}"].value}
allok &= result("status vocabulary readable", words <= {"PAID","DUE","PAST DUE","Upcoming",
                "ADJUSTED","OK","TIGHT","SHORTFALL"}, str(sorted(words)))

# --- ACTUAL / FORECAST present and distinguishing
tags = {cf[f"K{r}"].value for w in range(1, NW+1)
        for r in range(blk(w)["fs"], blk(w)["lv"]+1) if cf[f"K{r}"].value}
allok &= result("ACTUAL and FORECAST both present", tags == {"ACTUAL", "FORECAST"}, str(tags))

base_end = n(sm["I25"].value)

# --- overrides: amount, then removal
t = "/tmp/t_ovr.xlsx"; shutil.copy(src, t)
w = load_workbook(t); o = w["Overrides"]
o["B9"], o["C9"], o["E9"] = "Spectrum", dt.date(2026, 9, 21), 100.00
w.save(t); recalc(t)
after = n(load_workbook(t, data_only=True)["Summary"]["I25"].value)
allok &= result("override amount applies", abs(after - (base_end + 71.10)) < 0.02,
                f"{base_end:.2f} -> {after:.2f}")
w = load_workbook(t)
for c in "BCDEF": w["Overrides"][f"{c}9"] = None
w.save(t); recalc(t)
back = n(load_workbook(t, data_only=True)["Summary"]["I25"].value)
allok &= result("removing the override reverts exactly", abs(back - base_end) < 0.005,
                f"{back:.2f}")

# --- overrides: date move
shutil.copy(src, t); w = load_workbook(t); o = w["Overrides"]
o["B9"], o["C9"], o["D9"] = "Spectrum", dt.date(2026, 9, 21), dt.date(2026, 9, 26)
w.save(t); recalc(t)
vv = load_workbook(t, data_only=True)["Cash Flow"]
moved = [vv[f"C{r}"].value for w2 in range(1, NW+1)
         for r in range(blk(w2)["fs"], blk(w2)["fs"]+SLOTS)
         if vv[f"B{r}"].value == "Spectrum"]
allok &= result("override date moves the occurrence",
                any(d and d.month == 9 and d.day == 26 for d in moved), str([str(d)[:10] for d in moved]))

# --- the historical-payment bug: change a FUTURE date, past payment must not move
shutil.copy(src, t); w = load_workbook(t); rr = w["Recurring"]
for r in range(5, 35):
    if rr[f"B{r}"].value == "Progressive Insurance":
        rr[f"F{r}"] = dt.date(2026, 9, 2)
w.save(t); recalc(t)
vv = load_workbook(t, data_only=True)["Cash Flow"]
bruno = [(vv[f"B{r}"].value, n(vv[f"G{r}"].value)) for w2 in range(1, NW+1)
         for r in range(blk(w2)["fs"], blk(w2)["fs"]+SLOTS)
         if vv[f"B{r}"].value == "Bruno's Restaurant (paycheck)"]
allok &= result("past payment stays with its own transaction after a future date change",
                bruno and abs(bruno[0][1] - 148.34) < 0.005,
                f"Bruno's first occurrence = ${bruno[0][1]:.2f}" if bruno else "not found")

print("\nRESULT:", "ALL CHECKS PASS" if allok else "SOMETHING FAILED")
