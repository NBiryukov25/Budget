#!/usr/bin/env python3
"""Add a "Week Cards" tab: every week as a compact block, three to a row.

The Summary tab is one line per week, which is dense but hard to read - you
can see the totals but not what makes them up. The Cash Flow tab has the
detail but each week is fifty-odd rows tall, so only one week fits on screen
at a time. This sits in between: one card per week showing the scheduled
items, the three subtotals, the low point and the ending balance, laid out in
a grid you can scan.

Everything is a formula pointing at Summary and Cash Flow, so the cards track
edits made anywhere else in the workbook. Nothing is typed in.

    python3 add_week_cards.py <in.xlsx> <out.xlsx>
"""

import argparse
import re

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as col

FONT = "Aptos Narrow"
SHEET = "Week Cards"

ACROSS = 3          # cards per row
CARD_H = 22         # rows a card occupies, gutter included
TOP = 6             # first card row
LEFT = 2            # first card column (B)

MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY_Z = '$#,##0.00;($#,##0.00);$0.00'
DATE_S = 'ddd m/d'

NAVY = "1F3864"
GREY = "595959"

F_TITLE = Font(name=FONT, size=16, bold=True, color=NAVY)
F_SUB = Font(name=FONT, size=10, italic=True, color=GREY)
F_HEAD = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_BODY = Font(name=FONT, size=10, color="000000")
F_DIM = Font(name=FONT, size=9, color=GREY)
F_LBL = Font(name=FONT, size=10, color=GREY)
F_SUM = Font(name=FONT, size=10, bold=True, color="000000")
F_END = Font(name=FONT, size=12, bold=True, color="FFFFFF")
F_NOTE = Font(name=FONT, size=9, italic=True, color=GREY)

HAIR = Side(style="hair", color="BFBFBF")
RULE = Border(top=Side(style="thin", color="D0D0D0"))
EDGE = Side(style="thin", color="AEAEAE")

L, R, C = Alignment(horizontal="left"), Alignment(horizontal="right"), Alignment(horizontal="center")
LV = Alignment(horizontal="left", vertical="center")


SLOT_RANK = re.compile(r"MATCH\((\d+),\s*Schedule!")


def geometry(wb):
    """Map each week to the Cash Flow rows holding its scheduled items.

    The number of variable rows per week has changed more than once, which
    moves every block, so nothing here is hardcoded. The week banners are
    themselves formulas and so cannot be read as text; the slot rows can,
    because each one pulls its occurrence by rank and the rank encodes the
    week - 1001 is week 1 slot 1, 13010 is week 13 slot 10.
    """
    cf = wb["Cash Flow"]
    slots = {}
    for r in range(1, cf.max_row + 1):
        v = cf[f"B{r}"].value
        if not isinstance(v, str):
            continue
        m = SLOT_RANK.search(v)
        if m:
            slots.setdefault(int(m.group(1)) // 1000, []).append(r)
    if not slots:
        raise SystemExit("could not find the week blocks on the Cash Flow tab")
    return {w: (min(rows), len(rows)) for w, rows in slots.items()}


def build(wb):
    geo = geometry(wb)
    n_weeks = max(geo)
    slots = max(n for _, n in geo.values())
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, wb.sheetnames.index("Summary") + 1)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    for c in range(ACROSS):
        b = LEFT + c * 4
        ws.column_dimensions[col(b)].width = 8       # when
        ws.column_dimensions[col(b + 1)].width = 26  # what
        ws.column_dimensions[col(b + 2)].width = 12  # how much
        ws.column_dimensions[col(b + 3)].width = 2   # gutter

    ws["B2"] = "THE WEEKS AHEAD"
    ws["B2"].font = F_TITLE
    ws["B3"] = "Red banner: the week dips below zero. Amber: it stays up but lands under your cushion."
    ws["B3"].font = F_SUB
    ws["B4"] = "Nothing here is typed in. Edit elsewhere and these follow."
    ws["B4"].font = F_NOTE

    for w in range(1, n_weeks + 1):
        band, c = divmod(w - 1, ACROSS)
        t = TOP + band * CARD_H
        b = LEFT + c * 4
        card(ws, w, t, b, geo[w][0], slots)

    ws.freeze_panes = "A6"
    ws.sheet_view.zoomScale = 90
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    return ws, n_weeks


def card(ws, w, t, b, fs, items):
    """One week's card, top-left corner at row t, column b."""
    sr = 12 + w                                   # its row on Summary
    when, what, amt = col(b), col(b + 1), col(b + 2)
    S, CF = "Summary", "'Cash Flow'"

    # --- banner: week number and date range, coloured by how the week goes
    ws.merge_cells(f"{when}{t}:{amt}{t}")
    ws[f"{when}{t}"] = (f'="WEEK "&{S}!$B{sr}&"    "&TEXT({S}!$C{sr},"mmm d")'
                        f'&" - "&TEXT({S}!$C{sr}+6,"mmm d")')
    ws[f"{when}{t}"].font = F_HEAD
    ws[f"{when}{t}"].alignment = LV
    ws[f"{when}{t}"].fill = PatternFill("solid", fgColor="2E75B6")
    ws.row_dimensions[t].height = 20

    # --- opening balance
    ws[f"{when}{t + 1}"] = "Starts with"
    ws[f"{when}{t + 1}"].font = F_LBL
    ws[f"{amt}{t + 1}"] = f"={S}!$H{sr}"
    ws[f"{amt}{t + 1}"].font = F_BODY
    ws[f"{amt}{t + 1}"].number_format = MONEY
    ws[f"{amt}{t + 1}"].alignment = R

    ws[f"{when}{t + 2}"] = "SCHEDULED THIS WEEK"
    ws[f"{when}{t + 2}"].font = F_DIM
    for cc in (when, what, amt):
        ws[f"{cc}{t + 2}"].border = RULE

    # --- the scheduled items, straight off Cash Flow in date order
    fi = t + 3
    for i in range(items):
        r, src = fi + i, fs + i
        blank = f'{CF}!$B{src}=""'
        ws[f"{when}{r}"] = f'=IF({blank},"",{CF}!$C{src})'
        ws[f"{when}{r}"].number_format = DATE_S
        ws[f"{when}{r}"].font = F_DIM
        ws[f"{what}{r}"] = f'=IF({blank},"",{CF}!$B{src})'
        ws[f"{what}{r}"].font = F_BODY
        ws[f"{what}{r}"].alignment = L
        # income counts up, everything else counts down
        ws[f"{amt}{r}"] = (f'=IF({blank},"",IF({CF}!$G{src}<>"",{CF}!$G{src},-{CF}!$H{src}))')
        ws[f"{amt}{r}"].number_format = MONEY_Z
        ws[f"{amt}{r}"].font = F_BODY
        ws[f"{amt}{r}"].alignment = R
        ws.row_dimensions[r].height = 13

    # a used line gets a hairline under it; an empty one stays invisible
    ws.conditional_formatting.add(
        f"{when}{fi}:{amt}{fi + items - 1}",
        FormulaRule(formula=[f'${what}{fi}<>""'],
                    border=Border(bottom=HAIR), stopIfTrue=False))

    # --- the three subtotals
    li = t + 3 + items
    for k, (label, f, fnt) in enumerate([
            ("Money in", f"={S}!$D{sr}", Font(name=FONT, size=10, bold=True, color="006100")),
            ("Bills", f"=-{S}!$E{sr}", Font(name=FONT, size=10, color="9C0006")),
            ("Everyday spending", f"=-{S}!$F{sr}", Font(name=FONT, size=10, color="9C0006"))]):
        r = li + k
        ws[f"{when}{r}"] = label
        ws[f"{when}{r}"].font = F_LBL
        ws[f"{amt}{r}"] = f
        ws[f"{amt}{r}"].font = fnt
        ws[f"{amt}{r}"].number_format = MONEY
        ws[f"{amt}{r}"].alignment = R
        if k == 0:
            for cc in (when, what, amt):
                ws[f"{cc}{r}"].border = RULE

    # --- the low point: the number that actually decides whether a week works
    r = li + 3
    ws[f"{when}{r}"] = "Lowest point"
    ws[f"{when}{r}"].font = F_LBL
    ws[f"{what}{r}"] = f'=IF({S}!$K{sr}="","",TEXT({S}!$K{sr},"ddd mmm d"))'
    ws[f"{what}{r}"].font = F_DIM
    ws[f"{what}{r}"].alignment = R
    ws[f"{amt}{r}"] = f"={S}!$J{sr}"
    ws[f"{amt}{r}"].font = F_SUM
    ws[f"{amt}{r}"].number_format = MONEY
    ws[f"{amt}{r}"].alignment = R

    # --- closing balance, the card's headline
    e = li + 4
    ws.merge_cells(f"{when}{e}:{what}{e}")
    ws[f"{when}{e}"] = "ENDS WITH"
    ws[f"{when}{e}"].font = F_END
    ws[f"{when}{e}"].alignment = LV
    ws[f"{amt}{e}"] = f"={S}!$I{sr}"
    ws[f"{amt}{e}"].font = F_END
    ws[f"{amt}{e}"].number_format = MONEY
    ws[f"{amt}{e}"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[e].height = 22

    ws.merge_cells(f"{when}{e + 1}:{amt}{e + 1}")
    ws[f"{when}{e + 1}"] = f'={S}!$L{sr}'
    ws[f"{when}{e + 1}"].font = F_DIM
    ws[f"{when}{e + 1}"].alignment = C

    # banner and footer both take their colour from the week's status
    for rng, dull in ((f"{when}{t}:{amt}{t}", "2E75B6"), (f"{when}{e}:{amt}{e}", "2E75B6")):
        for test, shade in ((f'={S}!$J{sr}<0', "C00000"),
                            (f'={S}!$J{sr}<MinCushion', "BF8F00"),
                            (f'={S}!$J{sr}>=MinCushion', "548235")):
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[test[1:]],
                                 fill=PatternFill("solid", fgColor=shade), stopIfTrue=True))
        ws[rng.split(":")[0]].fill = PatternFill("solid", fgColor=dull)

    # a light box around the whole card so the grid reads as separate blocks
    for r in range(t, e + 2):
        for cc in (when, what, amt):
            cur = ws[f"{cc}{r}"].border
            ws[f"{cc}{r}"].border = Border(
                left=EDGE if cc == when else cur.left,
                right=EDGE if cc == amt else cur.right,
                top=EDGE if r == t else cur.top,
                bottom=EDGE if r == e + 1 else cur.bottom)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src")
    ap.add_argument("out")
    a = ap.parse_args()

    wb = load_workbook(a.src)
    ws, n = build(wb)
    wb.save(a.out)
    print(f"{SHEET}: {n} cards, {ACROSS} across -> {a.out}")


if __name__ == "__main__":
    main()
