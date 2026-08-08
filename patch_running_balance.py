#!/usr/bin/env python3
"""Patch an in-use workbook so the running balance follows date order.

The week blocks originally applied variable expenses in row order, which put
anything typed in the WEEKLY VARIABLE EXPENSES section after every recurring
transaction regardless of when it actually happened. A $6 expense on a Saturday
therefore landed after the following Wednesday's paycheck and the sheet
under-reported the mid-week low.

This rewrites only the ordering helpers, the Running Balance formulas and the
week-totals note. Everything the user has typed - beginning balance, Amount
Paid / Received entries, variable expense rows - is left untouched.

    python3 patch_running_balance.py <in.xlsx> <out.xlsx>
"""

import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

N_WEEKS, SLOTS, VAR_ROWS = 13, 10, 6
BLOCK_H = SLOTS + VAR_ROWS + 5
FIRST_BLOCK = 13
FONT = "Arial"
DATE_F = "ddd mm/dd"
F_INPUT = Font(name=FONT, size=10, color="0000FF")
F_NOTE_B = Font(name=FONT, size=9, bold=True, color="1F3864")


def main(src, dst):
    wb = load_workbook(src)
    if "Cash Flow" not in wb.sheetnames:
        raise SystemExit("no 'Cash Flow' tab - is this the right workbook?")
    cf = wb["Cash Flow"]

    cf.column_dimensions["K"].width = 52
    for col in ("L", "M"):
        cf.column_dimensions[col].width = 10
        cf.column_dimensions[col].hidden = True

    patched = 0
    for w in range(1, N_WEEKS + 1):
        h = FIRST_BLOCK + BLOCK_H * (w - 1)
        fs, ls = h + 2, h + 1 + SLOTS
        var_hdr = h + 2 + SLOTS
        fv, lv = h + 3 + SLOTS, h + 2 + SLOTS + VAR_ROWS
        tot = h + 3 + SLOTS + VAR_ROWS
        week_end_offset = 7 * (w - 1) + 6

        cf[f"B{var_hdr}"] = ("WEEKLY VARIABLE EXPENSES — type a description, THE DATE, "
                             "and the amount (groceries, gas, eating out, anything)")

        for r in range(fs, lv + 1):
            if r == var_hdr:
                continue
            is_var = r >= fv
            if is_var:
                cf[f"L{r}"] = (f'=IF($B{r}="","",IF($C{r}="",'
                               f'$E$6+{week_end_offset},$C{r}))')
                cf[f"C{r}"].number_format = DATE_F
                cf[f"C{r}"].font = F_INPUT
                cf[f"C{r}"].alignment = Alignment(horizontal="center")
            else:
                cf[f"L{r}"] = f'=IF($B{r}="","",$C{r})'
            cf[f"M{r}"] = (f'=IF($L{r}="","",COUNTIFS($L${fs}:$L${lv},"<"&$L{r})'
                           f'+COUNTIFS($L${fs}:$L{r},$L{r}))')
            cf[f"I{r}"] = (f'=IF($B{r}="","",$I${h}'
                           f'+SUMIFS($G${fs}:$G${lv},$M${fs}:$M${lv},"<="&$M{r})'
                           f'-SUMIFS($H${fs}:$H${lv},$M${fs}:$M${lv},"<="&$M{r}))')
            patched += 1

        cf[f"K{tot}"] = (
            f'=IF(COUNT($I${fs}:$I${lv})=0,"",'
            f'"Lowest point this week: "&TEXT(MIN($I${fs}:$I${lv}),"$#,##0.00;-$#,##0.00")'
            f'&" on "&TEXT(INDEX($C${fs}:$C${lv},MATCH(MIN($I${fs}:$I${lv}),'
            f'$I${fs}:$I${lv},0)),"ddd mmm d")&"     (in "&TEXT($G${tot},"$#,##0.00")'
            f'&", out "&TEXT($H${tot},"$#,##0.00")&")")')
        cf[f"K{tot}"].font = F_NOTE_B

    wb.save(dst)
    print(f"patched {patched} rows across {N_WEEKS} weeks -> {dst}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
