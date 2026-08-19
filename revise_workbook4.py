#!/usr/bin/env python3
"""Fourth pass: tidy the dashboard layout so it reads in one glance."""
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FONT = "Arial"
F_NOTE = Font(name=FONT, size=9, italic=True, color="595959")
FILL_CARD = PatternFill("solid", fgColor="F2F6FB")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ap = argparse.ArgumentParser(); ap.add_argument("src"); ap.add_argument("out")
a = ap.parse_args()
wb = load_workbook(a.src); sm = wb["Summary"]

# close the gaps: each card is label + one wide value, no orphan column
for r in range(4, 9):
    if f"D{r}:E{r}" not in [str(m) for m in sm.merged_cells.ranges]:
        try: sm.merge_cells(f"D{r}:E{r}")
        except Exception: pass
    sm[f"E{r}"].fill = FILL_CARD; sm[f"E{r}"].border = BOX
for r in (4, 5):
    try: sm.merge_cells(f"H{r}:J{r}")
    except Exception: pass
    for c in "IJ":
        sm[f"{c}{r}"].fill = FILL_CARD; sm[f"{c}{r}"].border = BOX
sm["D8"].alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)

# the explanatory line gets its own full-width row, clear of the chart
for m in [str(x) for x in sm.merged_cells.ranges if x.min_row == 10]:
    sm.unmerge_cells(m)
sm["B10"] = ("Cash available now is the beginning balance plus every transaction dated today "
             "or earlier. Lowest before then is the worst point between now and the next money "
             "in. Everyday Spending includes the top-up for weeks not yet logged.")
sm["B10"].font = F_NOTE
sm["B10"].alignment = Alignment(vertical="center", wrap_text=True)
sm.merge_cells("B10:L10")
sm.row_dimensions[10].height = 26

# move the chart clear of the table and let the print area reach it
if sm._charts:
    ch = sm._charts[0]
    ch.anchor = "B29"
    ch.height, ch.width = 9.5, 27
sm.print_area = "A1:L50"
wb.save(a.out)
print("stage 4 saved ->", a.out)
