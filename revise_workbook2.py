#!/usr/bin/env python3
"""Second pass: Cash Flow labelling, Summary dashboard, chart, Instructions."""

import argparse
import re

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

SLOTS, FIRST_BLOCK, N_WEEKS = 10, 13, 13
REC_FIRST, REC_LAST, REC_HDR = 5, 34, 4
SCH_FIRST = 4
LOG_FIRST, LOG_LAST = 9, 308
SM_HDR, SM_FIRST, SM_LAST = 12, 13, 25
CUR, PLOG, SCH = "'Cash Flow'", "'Paid Log'", "Schedule"

FONT = "Arial"
F_TITLE = Font(name=FONT, size=16, bold=True, color="1F3864")
F_SUB = Font(name=FONT, size=10, italic=True, color="595959")
F_H1 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_H2 = Font(name=FONT, size=10, bold=True, color="1F3864")
F_BODY = Font(name=FONT, size=10, color="000000")
F_INPUT = Font(name=FONT, size=10, color="0000FF")
F_BOLD = Font(name=FONT, size=10, bold=True, color="000000")
F_NOTE = Font(name=FONT, size=9, italic=True, color="595959")
F_KPI = Font(name=FONT, size=13, bold=True, color="1F3864")
F_LAB = Font(name=FONT, size=9, bold=True, color="595959")

FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_SUB = PatternFill("solid", fgColor="D9E2F3")
FILL_IN = PatternFill("solid", fgColor="FFFF00")
FILL_CARD = PatternFill("solid", fgColor="F2F6FB")
FILL_BAND = PatternFill("solid", fgColor="F2F2F2")

CHIP = {"good": ("006100", "C6EFCE"), "warn": ("9C5700", "FFEB9C"),
        "bad": ("9C0006", "FFC7CE"), "info": ("1F4E78", "DDEBF7"),
        "neutral": ("3F3F76", "EDEDED")}

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '$#,##0.00;($#,##0.00);"–"'
MONEY_B = '$#,##0.00;[Red]($#,##0.00);"–"'
DATE_F, DATE_L = "ddd mm/dd", "ddd mmm d, yyyy"
UNLOCKED, LOCKED = Protection(locked=False), Protection(locked=True)


def put(ws, ref, value, font=F_BODY, fmt=None, fill=None, align=None,
        border=None, wrap=False, unlocked=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if border: c.border = border
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.protection = UNLOCKED if unlocked else LOCKED
    return c


def chips(ws, rng, mapping):
    first = rng.split(":")[0]
    col = re.match(r"\$?([A-Z]+)", first).group(1)
    row = re.search(r"(\d+)", first).group(1)
    for word, kind in mapping.items():
        ink, ground = CHIP[kind]
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f'${col}{row}="{word}"'],
                             font=Font(name=FONT, size=9, bold=True, color=ink),
                             fill=PatternFill("solid", fgColor=ground)))


def detect_var_rows(cf):
    for vr in range(4, 41):
        h2 = FIRST_BLOCK + (SLOTS + vr + 5)
        v = cf[f"B{h2}"].value
        if isinstance(v, str) and v.startswith('="WEEK '):
            return vr
    raise SystemExit("cannot determine block height")


def block(w, vr):
    h = FIRST_BLOCK + (SLOTS + vr + 5) * (w - 1)
    return {"head": h, "cols": h + 1, "fs": h + 2, "ls": h + 1 + SLOTS,
            "vhdr": h + 2 + SLOTS, "fv": h + 3 + SLOTS,
            "lv": h + 2 + SLOTS + vr, "tot": h + 3 + SLOTS + vr}


def main():
    ap = argparse.ArgumentParser(); ap.add_argument("src"); ap.add_argument("out")
    a = ap.parse_args()
    wb = load_workbook(a.src)
    cf, rec, sm, sch, ins = (wb["Cash Flow"], wb["Recurring"], wb["Summary"],
                             wb["Schedule"], wb["Instructions"])
    VR = detect_var_rows(cf)
    SCH_LAST = SCH_FIRST
    while sch[f"A{SCH_LAST}"].value is not None:
        SCH_LAST += 1
    SCH_LAST -= 1

    LOG_D = f"{PLOG}!$B${LOG_FIRST}:$B${LOG_LAST}"
    LOG_N = f"{PLOG}!$C${LOG_FIRST}:$C${LOG_LAST}"
    SCH_C = f"{SCH}!$C${SCH_FIRST}:$C${SCH_LAST}"
    SCH_E = f"{SCH}!$E${SCH_FIRST}:$E${SCH_LAST}"
    SCH_M = f"{SCH}!$M${SCH_FIRST}:$M${SCH_LAST}"
    SCH_NK = f"{SCH}!$N${SCH_FIRST}:$N${SCH_LAST}"

    # ================================ 9 + 2 + 14 · Cash Flow: labels and chips
    cf.column_dimensions["K"].width = 13
    cf.column_dimensions["N"].width = 54
    status_rngs, tag_rngs = [], []
    for w in range(1, N_WEEKS + 1):
        b = block(w, VR)
        h, cols, fs, ls, fv, lv, tot = (b["head"], b["cols"], b["fs"], b["ls"],
                                        b["fv"], b["lv"], b["tot"])
        # the long notes move out of K so K can be a narrow labelling column
        for row in (h, tot):
            if cf[f"K{row}"].value is not None:
                src = cf[f"K{row}"]
                cf[f"N{row}"] = src.value
                cf[f"N{row}"].font = Font(name=FONT, size=src.font.sz or 9,
                                          bold=bool(src.font.b),
                                          italic=bool(src.font.i),
                                          color=(src.font.color.rgb
                                                 if src.font.color and
                                                 isinstance(src.font.color.rgb, str)
                                                 else "595959"))
                cf[f"K{row}"] = None
        put(cf, f"K{cols}", "Actual / Forecast", F_H2, fill=FILL_SUB,
            align="center", border=BOX, wrap=True)
        hits = f'COUNTIFS({LOG_D},$C%s,{LOG_N},$B%s)'
        for r in range(fs, ls + 1):
            cf[f"K{r}"] = (f'=IF($B{r}="","",IF(OR($F{r}<>"",{hits % (r, r)}>0),'
                           f'"ACTUAL","FORECAST"))')
            cf[f"K{r}"].font = Font(name=FONT, size=9, bold=True)
            cf[f"K{r}"].alignment = Alignment(horizontal="center")
            cf[f"F{r}"].protection = UNLOCKED
        for r in range(fv, lv):
            cf[f"K{r}"] = f'=IF($B{r}="","","ACTUAL")'
            cf[f"K{r}"].font = Font(name=FONT, size=9, bold=True)
            cf[f"K{r}"].alignment = Alignment(horizontal="center")
            for col in "BCE":
                cf[f"{col}{r}"].protection = UNLOCKED
        cf[f"K{lv}"] = '="FORECAST"'
        cf[f"K{lv}"].font = Font(name=FONT, size=9, bold=True)
        cf[f"K{lv}"].alignment = Alignment(horizontal="center")
        status_rngs.append(f"J{fs}:J{lv}")
        status_rngs.append(f"J{tot}:J{tot}")
        tag_rngs.append(f"K{fs}:K{lv}")
    for rng in status_rngs:
        chips(cf, rng, {"PAID": "good", "OK": "good", "ADJUSTED": "info",
                        "DUE": "warn", "TIGHT": "warn", "Upcoming": "neutral",
                        "PAST DUE": "bad", "SHORTFALL": "bad"})
    for rng in tag_rngs:
        chips(cf, rng, {"ACTUAL": "info", "FORECAST": "neutral"})
    for ref in ("E5", "E6", "E7", "E8"):
        cf[ref].protection = UNLOCKED

    # 11 · a short legend so the input convention is stated, not inferred
    put(cf, "B12", "Blue text on yellow = you type it   ·   ACTUAL = already happened   ·   "
                   "FORECAST = still expected   ·   everything else is calculated",
        F_NOTE)

    # ==================================================== 7 + 15 · the dashboard
    # unmerge before clearing: a MergedCell's value is read-only
    for mc in [str(m) for m in sm.merged_cells.ranges
               if m.min_row >= 3 and m.max_row <= 11]:
        sm.unmerge_cells(mc)
    for row in range(3, 12):
        for col in "BCDEFGHIJKL":
            c = sm[f"{col}{row}"]
            c.value = None
            c.fill = PatternFill()

    # hidden scalars driving the dashboard
    put(sm, "N4", f'=IF(COUNT({SCH_NK})=0,"",MIN({SCH_NK}))', F_NOTE)
    put(sm, "N5", f'=IF(COUNT({SCH_M})=0,"",MIN({SCH_M}))', F_NOTE)
    put(sm, "N6", '=IF($N$4="","",INT($N$4/100000))', F_NOTE, fmt=DATE_L)
    put(sm, "N7", f'=IF($N$4="","",INDEX({SCH_C},MATCH($N$4,{SCH_NK},0)))', F_NOTE)
    put(sm, "N8", f'=IF($N$4="","",INDEX({SCH_E},MATCH($N$4,{SCH_NK},0)))', F_NOTE, fmt=MONEY)
    put(sm, "N9", '=IF($N$5="","",INT($N$5/100000))', F_NOTE, fmt=DATE_L)
    put(sm, "N10", f'=IF($N$5="","",INDEX({SCH_C},MATCH($N$5,{SCH_M},0)))', F_NOTE)
    put(sm, "N11", f'=IF($N$5="","",INDEX({SCH_E},MATCH($N$5,{SCH_M},0)))', F_NOTE, fmt=MONEY)

    put(sm, "B3", "AT A GLANCE", F_H2)
    left = [
        ("Cash available now",  '=BeginningBalance+SUM($N${0}:$N${1})-SUM($O${0}:$O${1})'
                                .format(SM_FIRST, SM_LAST), MONEY_B),
        ("Lowest before then",  '=IF($N$4="","",MIN($D$4,MIN($P${0}:$P${1})))'
                                .format(SM_FIRST, SM_LAST), MONEY_B),
        ("Lowest in projection", f'=MIN($J${SM_FIRST}:$J${SM_LAST})', MONEY_B),
        ("Ending balance",      f'=$I${SM_LAST}', MONEY_B),
    ]
    for i, (lab, formula, fmt) in enumerate(left):
        r = 4 + i
        put(sm, f"B{r}", lab, F_LAB, fill=FILL_CARD, align="left", border=BOX)
        sm.merge_cells(f"B{r}:C{r}")
        sm[f"C{r}"].fill = FILL_CARD; sm[f"C{r}"].border = BOX
        put(sm, f"D{r}", formula, F_KPI, fmt=fmt, fill=FILL_CARD,
            align="center", border=BOX)
    put(sm, "B8", "Next money in", F_LAB, fill=FILL_CARD, align="left", border=BOX)
    sm.merge_cells("B8:C8")
    sm["C8"].fill = FILL_CARD; sm["C8"].border = BOX
    put(sm, "D8", '=IF($N$4="","none scheduled",$N$7&"  "&TEXT($N$8,"$#,##0.00")&'
                  '"  on "&TEXT($N$6,"ddd mmm d"))', F_BOLD, fill=FILL_CARD,
        align="left", border=BOX)
    sm.merge_cells("D8:E8")
    sm["E8"].fill = FILL_CARD; sm["E8"].border = BOX

    right = [
        ("Minimum cash cushion", "=MinCushion", MONEY),
        ("Short of cushion by",
         f'=MAX(0,MinCushion-MIN($J${SM_FIRST}:$J${SM_LAST}))', MONEY),
        ("Periods going negative",
         f'=COUNTIF($J${SM_FIRST}:$J${SM_LAST},"<0")&" of {N_WEEKS}"', None),
        ("Next major bill",
         '=IF($N$5="","none over $100",$N$10&"  "&TEXT($N$11,"$#,##0.00")&'
         '"  on "&TEXT($N$9,"ddd mmm d"))', None),
    ]
    for i, (lab, formula, fmt) in enumerate(right):
        r = 4 + i
        put(sm, f"F{r}", lab, F_LAB, fill=FILL_CARD, align="left", border=BOX)
        sm.merge_cells(f"F{r}:G{r}")
        sm[f"G{r}"].fill = FILL_CARD; sm[f"G{r}"].border = BOX
        put(sm, f"H{r}", formula, F_BOLD if i >= 2 else F_KPI, fmt=fmt,
            fill=FILL_CARD, align="center" if i < 2 else "left", border=BOX)
        if i >= 2:
            sm.merge_cells(f"H{r}:J{r}")
            for cc in "IJ":
                sm[f"{cc}{r}"].fill = FILL_CARD; sm[f"{cc}{r}"].border = BOX

    put(sm, "F8", "Needs attention", F_LAB, fill=FILL_CARD, align="left", border=BOX)
    sm.merge_cells("F8:G8")
    sm["G8"].fill = FILL_CARD; sm["G8"].border = BOX
    put(sm, "H8",
        f'=IF(COUNTIF(Recurring!$K${REC_FIRST}:$K${REC_LAST},"!*")=0,'
        f'"All recurring items complete",'
        f'"! "&COUNTIF(Recurring!$K${REC_FIRST}:$K${REC_LAST},"!*")&'
        f'" recurring item(s) need attention — see the Recurring tab")',
        F_BOLD, fill=FILL_CARD, align="left", border=BOX)
    sm.merge_cells("H8:L8")
    for cc in "IJKL":
        sm[f"{cc}8"].fill = FILL_CARD; sm[f"{cc}8"].border = BOX
    sm.conditional_formatting.add("H8:L8", FormulaRule(
        formula=['LEFT($H$8,1)="!"'],
        font=Font(name=FONT, size=10, bold=True, color=CHIP["bad"][0]),
        fill=PatternFill("solid", fgColor=CHIP["bad"][1])))
    for ref in ("D4", "D5", "D6", "D7", "H5"):
        sm.conditional_formatting.add(ref, FormulaRule(
            formula=[f'AND(ISNUMBER({ref}),{ref}<0)'],
            font=Font(name=FONT, size=13, bold=True, color=CHIP["bad"][0])))

    put(sm, "B10", "Cash available now is the beginning balance plus every transaction "
                   "dated today or earlier. Lowest before then is the worst point between "
                   "now and the next money in.", F_NOTE)

    # per-week helpers behind the dashboard and the chart
    for col, w in (("N", 12), ("O", 12), ("P", 12), ("Q", 10), ("R", 10)):
        sm.column_dimensions[col].width = w
    put(sm, f"Q{SM_HDR}", "Zero", F_LAB, align="center")
    put(sm, f"R{SM_HDR}", "Cushion", F_LAB, align="center")
    for w in range(1, N_WEEKS + 1):
        r = SM_HDR + w
        b = block(w, VR)
        cr = f"{CUR}!$C${b['fs']}:$C${b['lv']}"
        put(sm, f"N{r}", f'=SUMIFS({CUR}!$G${b["fs"]}:$G${b["lv"]},{cr},"<="&TODAY())',
            F_NOTE, fmt=MONEY)
        put(sm, f"O{r}", f'=SUMIFS({CUR}!$H${b["fs"]}:$H${b["lv"]},{cr},"<="&TODAY())',
            F_NOTE, fmt=MONEY)
        put(sm, f"P{r}",
            f'=IF($N$4="","",IF(COUNTIFS({cr},">="&TODAY(),{cr},"<="&$N$6)=0,"",'
            f'_xlfn.MINIFS({CUR}!$I${b["fs"]}:$I${b["lv"]},{cr},">="&TODAY(),'
            f'{cr},"<="&$N$6)))', F_NOTE, fmt=MONEY)
        put(sm, f"Q{r}", 0, F_NOTE, fmt=MONEY)
        put(sm, f"R{r}", "=MinCushion", F_NOTE, fmt=MONEY)
    for col in "NOP":
        sm.column_dimensions[col].hidden = True

    chips(sm, f"L{SM_FIRST}:L{SM_LAST}",
          {"OK": "good", "TIGHT": "warn", "SHORTFALL": "bad"})

    # ================================================ 8 · projected balance chart
    ch = LineChart()
    ch.title = "Projected balance by week"
    ch.style = 2
    ch.height, ch.width = 8.5, 26
    ch.y_axis.title = "Balance"
    ch.y_axis.numFmt = '$#,##0'
    data = Reference(sm, min_col=9, max_col=10, min_row=SM_HDR, max_row=SM_LAST)
    zero = Reference(sm, min_col=17, max_col=18, min_row=SM_HDR, max_row=SM_LAST)
    cats = Reference(sm, min_col=3, min_row=SM_FIRST, max_row=SM_LAST)
    ch.add_data(data, titles_from_data=True)
    ch.add_data(zero, titles_from_data=True)
    ch.set_categories(cats)
    palette = ["5B9BD5", "E3A93C", "F0605D", "9AA6B8"]
    for s, colour in zip(ch.series, palette):
        s.graphicalProperties.line.solidFill = colour
        s.graphicalProperties.line.width = 22000
        s.smooth = False
    for s in ch.series[2:]:
        s.graphicalProperties.line.dashStyle = "dash"
        s.graphicalProperties.line.width = 14000
    sm.add_chart(ch, "B28")

    # ============================================ 3 · Instructions, made current
    for mc in [str(m) for m in ins.merged_cells.ranges]:
        ins.unmerge_cells(mc)
    for row in ins.iter_rows(min_row=1, max_row=200, max_col=8):
        for c in row:
            c.value = None
            c.fill = PatternFill()
            c.border = Border()
    for col, w in {"A": 3, "B": 26, "C": 96, "D": 3}.items():
        ins.column_dimensions[col].width = w

    put(ins, "B2", "WEEKLY CASH FLOW SURVIVAL WORKSHEET", F_TITLE)
    put(ins, "B3", '="Currently projecting "&TEXT(StartDate,"mmm d, yyyy")&" to "'
                   f'&TEXT(StartDate+{N_WEEKS * 7 - 1},"mmm d, yyyy")&" from a beginning '
                   'balance of "&TEXT(BeginningBalance,"$#,##0.00")&"."', F_SUB)

    r = 5
    def section(title):
        nonlocal r
        put(ins, f"B{r}", title, F_H2); r += 1

    def line(label, body, h=30):
        nonlocal r
        put(ins, f"B{r}", label, F_BOLD)
        put(ins, f"C{r}", body, F_BODY, wrap=True)
        ins.row_dimensions[r].height = h
        r += 1

    section("THE FIVE PLACES YOU TYPE")
    line("Cash Flow, E5–E8", '="Beginning balance "&TEXT(BeginningBalance,"$#,##0.00")&'
                             '", start date "&TEXT(StartDate,"mmm d, yyyy")&", cushion "&'
                             'TEXT(MinCushion,"$#,##0.00")&", expected everyday spending "&'
                             'TEXT(WeeklySpend,"$#,##0.00")&" a week. Change any of these and '
                             'the whole workbook follows."', 34)
    line("Recurring", "Every bill and paycheck that repeats. Edit a row, add one in the first "
                      "empty row, or set Active to No to retire it without losing the record. "
                      "The Needs Attention column tells you when something is incomplete.")
    line("Cash Flow, week blocks", "The variable-expense rows in each week: description, date, "
                                   "amount. Everything you actually spend goes here.")
    line("Paid Log", "What a scheduled bill or paycheck really came to. Keyed to the "
                     "transaction, so it stays put however the schedule re-sorts.")
    line("Overrides", "A one-off different date or amount for a single occurrence. Delete the "
                      "row to go back to normal.")

    r += 1
    section("HOW TO READ IT")
    line("ACTUAL vs FORECAST", "Every transaction row is labelled. ACTUAL has already "
                               "happened or been logged; FORECAST is still expected.")
    line("Status", "PAID and OK are green. DUE and TIGHT are amber. PAST DUE and SHORTFALL "
                   "are red. TIGHT means above zero but under your cushion.")
    line("Lowest In Week", "The worst moment inside a week, not just where it ends. A week can "
                           "close healthy having gone under midweek, so this is what Status "
                           "is judged on.")
    line("Weekly Surplus", "The week standing alone — money in less all spending, nothing "
                           "carried in. Carried Forward is shown separately.")
    line("Colours", "Blue text on yellow means you type it. Everything else is calculated.")

    r += 1
    section("RULES WORTH KNOWING")
    line("Next Due Date", "The next occurrence you have NOT paid. A date in the past is fine "
                          "— it shows in the first period as PAST DUE. Only move it forward "
                          "once the money has actually moved.")
    line("Dates", "Type 8/21/2026, not 821. A bare 821 is the number 821, which Excel reads "
                  "as a date in 1902 and the row disappears. The Needs Attention column "
                  "catches this.")
    line("Start date", '="Any weekday works. Yours begins on a "&TEXT(StartDate,"dddd")&'
                       '", so every period runs "&TEXT(StartDate,"ddd")&" to "&'
                       'TEXT(StartDate+6,"ddd")&"."')
    line("Everyday spending", "Weeks with nothing logged are topped up to your expected "
                              "weekly figure, so an untouched future week is never forecast "
                              "at zero. Log more than it and the top-up falls away.")
    line("Schedule tab", "The engine that turns recurring rules into dated occurrences. Safe "
                         "to look at, protected against typing.")

    r += 1
    section("ASSUMPTIONS ON RECORD")
    for body in [
        "Recurring items were transcribed from the user's budget-app screenshots, August 2026. "
        "Nothing was invented.",
        "Amounts are positive numbers; the Type column decides the direction.",
        '="Compunnel is entered at "&TEXT(INDEX(Recurring!$D:$D,MATCH("Compunnel*",'
        'Recurring!$B:$B,0)),"$#,##0.00")&", the low end of a paycheck that varies by about '
        '$45. A survival sheet plans on the smaller number."',
        "Microsoft is active with no amount. It is flagged rather than guessed — put the real "
        "figure on the Recurring tab.",
        "The Central Maine Power payment plan still repeats monthly alongside the $163.50 "
        "bill. If the plan ends when normal billing resumes, set it to Active = No.",
    ]:
        put(ins, f"B{r}", "•", F_BODY, align="center")
        put(ins, f"C{r}", body, F_NOTE, wrap=True)
        ins.row_dimensions[r].height = 26
        r += 1

    ins.print_area = f"A1:C{r}"
    wb.save(a.out)
    print(f"stage 2 saved -> {a.out}  (block {VR} var rows, schedule to {SCH_LAST})")


if __name__ == "__main__":
    main()
