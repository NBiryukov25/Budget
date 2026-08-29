#!/usr/bin/env python3
"""Replace the Week Cards tab with an everyday-spending tracker.

The weekly target only works if you can see where you are against it before
the week is over. This adds a tab that answers three questions: what is left
to spend this week, what that leaves per day, and which categories the money
actually goes to.

Everything reads off the variable-expense rows already in Cash Flow, so
expenses keep being entered exactly where they are now - by hand or by the
importer. Nothing here is typed in twice.

Classification lives in a hidden column on Cash Flow rather than here, so the
totals are one SUMIFS each instead of a keyword search repeated per week.

    python3 add_spending_tracker.py <in.xlsx> <out.xlsx>
"""

import argparse
import re

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as col

FONT = "Aptos Narrow"
SHEET = "Everyday Spending"
CAT_COL = "N"                      # hidden helper on Cash Flow
CF = "'Cash Flow'"

MONEY = '$#,##0.00;($#,##0.00);"-"'
NAVY, GREY, RED, GREEN, AMBER = "1F3864", "595959", "9C0006", "006100", "BF8F00"

F_TITLE = Font(name=FONT, size=16, bold=True, color=NAVY)
F_SUB = Font(name=FONT, size=10, italic=True, color=GREY)
F_H = Font(name=FONT, size=10, bold=True, color="FFFFFF")
F_B = Font(name=FONT, size=10, color="000000")
F_LBL = Font(name=FONT, size=10, color=GREY)
F_BIG = Font(name=FONT, size=20, bold=True, color=NAVY)
F_BIG2 = Font(name=FONT, size=14, bold=True, color="000000")
F_NOTE = Font(name=FONT, size=9, italic=True, color=GREY)

FILL_H = PatternFill("solid", fgColor="2E75B6")
FILL_PANEL = PatternFill("solid", fgColor="EAF1FA")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
R, C, L = (Alignment(horizontal=h) for h in ("right", "center", "left"))

# Drawn from the merchants that actually appear on the statements. A name that
# matches nothing lands in "Everything else", which is the cue to add it here.
CATS = [
    ("Rent paid in cash",   ["cash back"]),
    ("Groceries",           ["hannaford", "market basket", "shaws", "westbrook market"]),
    ("Convenience & gas",   ["cumberland", "7-eleven", "7 eleven", "circle k", "first stop",
                             "buxton", "big apple", "shell", "exxon", "dollar tree",
                             "walgreens", "advance auto"]),
    ("Cash withdrawals",    ["cash withdrawal", "atm"]),
    ("Eating out",          ["subway", "dunkin", "wendy", "kfc", "amatos", "idexx",
                             "coff", "portland smoke"]),
    ("Online & apps",       ["eneba", "kalshi", "venice", "telegram", "steam", "mage space",
                             "render", "buy me a coffee", "openai", "obsidian", "claude",
                             "apple", "google", "spotify", "microsoft", "netflix"]),
    ("Money transfers",     ["remitly", "smart glocal", "rombalski", "intrnet"]),
    ("Loan repayments",     ["earnin", "repayment", "upstart"]),
]
# neither is living costs: one is rent that merely left the account as cash,
# the other is debt. Both belong in the totals but not in what you compare
# against the weekly target.
NOT_LIVING = ("Rent paid in cash", "Loan repayments")
OTHER = "Everything else"
PLANNED = "(not yet spent)"


def geometry(wb):
    """Variable-expense rows, read off the workbook rather than assumed."""
    cf = wb["Cash Flow"]
    pat = re.compile(r"MATCH\((\d+),\s*Schedule!")
    slots = {}
    for r in range(1, cf.max_row + 1):
        v = cf[f"B{r}"].value
        if isinstance(v, str):
            m = pat.search(v)
            if m:
                slots.setdefault(int(m.group(1)) // 1000, []).append(r)
    weeks = sorted(slots)
    n_slots = max(len(v) for v in slots.values())
    first = min(slots[weeks[0]]) + n_slots + 1          # past the slot header
    block = min(slots[weeks[1]]) - min(slots[weeks[0]])
    var = block - n_slots - 5
    return {"weeks": weeks, "block": block, "slots": n_slots,
            "first_var": first, "var": var,
            "last_var": min(slots[weeks[-1]]) + n_slots + var}


def classify(cell):
    """Nested IFs over the keyword lists - one label per row, or blank."""
    out = f'IF({cell}="","",IF(LEFT({cell},17)="Everyday spending","{PLANNED}",'
    close = 2
    for name, keys in CATS:
        tests = ",".join(f'ISNUMBER(SEARCH("{k}",{cell}))' for k in keys)
        out += f'IF(OR({tests}),"{name}",'
        close += 1
    out += f'"{OTHER}"' + ")" * close
    return "=" + out


def tag_cash_flow(wb, geo):
    """Label every variable row, so the tracker needs one SUMIFS per category."""
    cf = wb["Cash Flow"]
    n = 0
    for w in geo["weeks"]:
        fv = geo["first_var"] + geo["block"] * (w - 1)
        for r in range(fv, fv + geo["var"] + 1):
            cf[f"{CAT_COL}{r}"] = classify(f"$B{r}")
            cf[f"{CAT_COL}{r}"].font = F_NOTE
            n += 1
    cf.column_dimensions[CAT_COL].hidden = True
    return n


def build(wb, geo):
    if "Week Cards" in wb.sheetnames:
        del wb["Week Cards"]
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, wb.sheetnames.index("Summary") + 1)
    ws.sheet_view.showGridLines = False

    lo, hi = geo["first_var"], geo["last_var"]
    AMT = f"{CF}!$E${lo}:$E${hi}"
    CAT = f"{CF}!${CAT_COL}${lo}:${CAT_COL}${hi}"
    DAY = f"{CF}!$C${lo}:$C${hi}"
    # non-blank category excludes the scheduled-bill rows; the top-up row is a
    # placeholder for money not yet spent and would otherwise peg every week at
    # exactly the target
    REAL = f'{CAT},"<>",{CAT},"<>{PLANNED}"'

    for c, wdt in zip("ABCDEFGH", (2, 26, 15, 13, 13, 13, 13, 15)):
        ws.column_dimensions[c].width = wdt

    ws["B2"] = "EVERYDAY SPENDING"
    ws["B2"].font = F_TITLE
    ws["B3"] = ("Groceries, gas, cash, eating out - everything that is not a scheduled bill. "
                "Keep entering it on the Cash Flow tab; this just keeps score.")
    ws["B3"].font = F_SUB

    # ---------------------------------------------------------------- this week
    ws.merge_cells("B5:H5")
    ws["B5"] = "THIS WEEK"
    ws["B5"].font = F_H
    ws["B5"].fill = FILL_H
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 20

    wk = f'MIN({len(geo["weeks"])},MAX(1,INT((TODAY()-StartDate)/7)+1))'
    ws["B6"] = "Week"
    ws["C6"] = f"={wk}"
    ws["B7"] = "Running"
    ws["C7"] = "=StartDate+($C$6-1)*7"
    ws["D7"] = '="to "&TEXT($C$7+6,"ddd mmm d")'
    for r in (6, 7):
        ws[f"B{r}"].font = F_LBL
        ws[f"C{r}"].font = F_B
    ws["C7"].number_format = "ddd mmm d"
    ws["D7"].font = F_LBL

    panel = [("Target", "=WeeklySpend", F_BIG2),
             ("Spent so far", f"=SUMIFS({AMT},{REAL},{DAY},\">=\"&$C$7,{DAY},\"<=\"&$C$7+6)", F_BIG2),
             ("Left to spend", "=$C$9-$C$10", F_BIG),
             ("Days left", "=MAX(0,$C$7+6-TODAY()+1)", F_BIG2),
             ("A day from here", '=IF($C$12<=0,"",MAX(0,$C$11/$C$12))', F_BIG)]
    for i, (label, f, fnt) in enumerate(panel):
        r = 9 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = F_LBL
        ws[f"C{r}"] = f
        ws[f"C{r}"].font = fnt
        ws[f"C{r}"].number_format = MONEY if label != "Days left" else "0"
        ws[f"C{r}"].alignment = R
        for c in "BC":
            ws[f"{c}{r}"].fill = FILL_PANEL
    ws.row_dimensions[11].height = 26
    ws.row_dimensions[13].height = 26

    ws["D11"] = ('=IF($C$11<0,"over by "&TEXT(-$C$11,"$#,##0.00"),'
                 'IF($C$12<=0,"week is done","still to go"))')
    ws["D11"].font = F_LBL
    ws["D13"] = ('=IF($C$12<=0,"",IF($C$11<=0,"nothing left",'
                 '"per day for the "&$C$12&" days left"))')
    ws["D13"].font = F_LBL
    ws.conditional_formatting.add("C11:C13", CellIsRule(
        operator="lessThan", formula=["0"], font=Font(name=FONT, size=20, bold=True, color=RED)))

    # ------------------------------------------------------------ week by week
    hdr = ["Week", "Beginning", "Target", "Spent", "Left", "A day", "Status"]
    for i, h in enumerate(hdr):
        c = col(2 + i)
        ws[f"{c}16"] = h
        ws[f"{c}16"].font = F_H
        ws[f"{c}16"].fill = FILL_H
        ws[f"{c}16"].alignment = C if i else L

    for i, w in enumerate(geo["weeks"]):
        r = 17 + i
        beg = f"StartDate+{w - 1}*7"
        spent = f'SUMIFS({AMT},{REAL},{DAY},">="&{beg},{DAY},"<="&{beg}+6)'
        vals = [w, f"={beg}", "=WeeklySpend", f"={spent}", f"=$D{r}-$E{r}",
                f'=IF($E{r}=0,"",$E{r}/7)',
                (f'=IF(TODAY()<{beg},"not yet",IF($E{r}=0,"nothing logged",'
                 f'IF($F{r}<0,"OVER","under")))')]
        for j, v in enumerate(vals):
            c = col(2 + j)
            ws[f"{c}{r}"] = v
            ws[f"{c}{r}"].font = F_B
            ws[f"{c}{r}"].border = BOX
            if j == 1:
                ws[f"{c}{r}"].number_format = "ddd mmm d"
            elif j in (2, 3, 4, 5):
                ws[f"{c}{r}"].number_format = MONEY
                ws[f"{c}{r}"].alignment = R
            else:
                ws[f"{c}{r}"].alignment = C
        ws.conditional_formatting.add(
            f"B{r}:H{r}", FormulaRule(formula=[f"$F{r}<0"],
                                      fill=PatternFill("solid", fgColor="FCE4D6"),
                                      stopIfTrue=False))

    last = 16 + len(geo["weeks"])
    ws[f"B{last + 1}"] = "Weeks logged"
    ws[f"B{last + 1}"].font = F_LBL
    ws[f"E{last + 1}"] = f"=SUM($E$17:$E${last})"
    ws[f"F{last + 1}"] = f"=SUM($F$17:$F${last})"
    for c in "EF":
        ws[f"{c}{last + 1}"].font = Font(name=FONT, size=10, bold=True)
        ws[f"{c}{last + 1}"].number_format = MONEY
        ws[f"{c}{last + 1}"].alignment = R

    # -------------------------------------------------------------- categories
    top = last + 4
    ws.merge_cells(f"B{top}:H{top}")
    ws[f"B{top}"] = "WHERE IT GOES"
    ws[f"B{top}"].font = F_H
    ws[f"B{top}"].fill = FILL_H
    ws[f"B{top}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[top].height = 20

    for i, h in enumerate(["Category", "This week", "Every week", "Share", "A week"]):
        c = col(2 + i)
        ws[f"{c}{top + 1}"] = h
        ws[f"{c}{top + 1}"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws[f"{c}{top + 1}"].alignment = C if i else L

    names = [n for n, _ in CATS] + [OTHER]
    f0 = top + 2
    for i, name in enumerate(names):
        r = f0 + i
        ws[f"B{r}"] = name
        ws[f"B{r}"].font = F_B
        ws[f"C{r}"] = (f'=SUMIFS({AMT},{CAT},$B{r},{DAY},">="&$C$7,{DAY},"<="&$C$7+6)')
        ws[f"D{r}"] = f"=SUMIFS({AMT},{CAT},$B{r})"
        ws[f"E{r}"] = f'=IF($D${f0 + len(names)}=0,"",$D{r}/$D${f0 + len(names)})'
        ws[f"F{r}"] = (f'=IF(COUNTIF($E$17:$E${last},">0")=0,"",'
                       f'$D{r}/COUNTIF($E$17:$E${last},">0"))')
        for c in "CDF":
            ws[f"{c}{r}"].number_format = MONEY
            ws[f"{c}{r}"].alignment = R
        ws[f"E{r}"].number_format = "0%"
        ws[f"E{r}"].alignment = R
        for c in "BCDEF":
            ws[f"{c}{r}"].font = F_B
            ws[f"{c}{r}"].border = BOX

    tr = f0 + len(names)
    ws[f"B{tr}"] = "All everyday spending"
    ws[f"C{tr}"] = f"=SUM($C${f0}:$C${tr - 1})"
    ws[f"D{tr}"] = f"=SUM($D${f0}:$D${tr - 1})"
    ws[f"F{tr}"] = f"=SUM($F${f0}:$F${tr - 1})"
    for c in "BCDF":
        ws[f"{c}{tr}"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws[f"{c}{tr}"].border = Border(top=Side(style="double", color=NAVY))
        if c != "B":
            ws[f"{c}{tr}"].number_format = MONEY
            ws[f"{c}{tr}"].alignment = R

    lr = tr + 1
    excl = "+".join(f"$D${f0 + names.index(x)}" for x in NOT_LIVING)
    ws[f"B{lr}"] = "Living costs only - compare this with the target"
    ws[f"C{lr}"] = f"=$C${tr}-" + "-".join(f"$C${f0 + names.index(x)}" for x in NOT_LIVING)
    ws[f"D{lr}"] = f"=$D${tr}-({excl})"
    ws[f"F{lr}"] = (f'=IF(COUNTIF($E$17:$E${last},">0")=0,"",'
                    f'$D{lr}/COUNTIF($E$17:$E${last},">0"))')
    for c in "BCDF":
        ws[f"{c}{lr}"].font = Font(name=FONT, size=11, bold=True, color=NAVY)
        if c != "B":
            ws[f"{c}{lr}"].number_format = MONEY
            ws[f"{c}{lr}"].alignment = R
    ws.conditional_formatting.add(f"F{lr}", CellIsRule(
        operator="greaterThan", formula=["WeeklySpend"],
        font=Font(name=FONT, size=11, bold=True, color=RED)))

    ws[f"B{tr + 3}"] = ('"Every week" counts only weeks with spending logged, so the per-week '
                        'figure is a real average rather than one dragged down by weeks that '
                        'have not happened. Anything landing in "Everything else" needs its '
                        'merchant adding to the category list.')
    ws[f"B{tr + 3}"].font = F_NOTE
    ws.freeze_panes = "A16"
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src")
    ap.add_argument("out")
    a = ap.parse_args()
    wb = load_workbook(a.src)
    geo = geometry(wb)
    n = tag_cash_flow(wb, geo)
    build(wb, geo)
    wb.save(a.out)
    print(f"{SHEET}: built, {n} rows classified in hidden column {CAT_COL} -> {a.out}")


if __name__ == "__main__":
    main()
