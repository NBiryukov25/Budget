#!/usr/bin/env python3
"""Build Weekly_Cash_Flow_Survival.xlsx.

Everything the workbook computes is a live Excel formula, so editing the
Recurring sheet or the beginning balance re-drives the whole model.

Layout
------
Instructions  how to use it, colour legend, assumptions
Recurring     the editable recurring-transaction table (30 slots)
Cash Flow     13 weekly blocks: beginning balance -> transactions -> ending balance
Summary       week-by-week rollup and the survival warnings
Schedule      formula engine that expands recurring rules into dated occurrences
"""

import datetime as dt

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "Weekly_Cash_Flow_Survival.xlsx"

FONT = "Arial"

# ---------------------------------------------------------------- dimensions
N_ITEMS = 30           # recurring-transaction slots
N_OCC = 16             # occurrences generated per recurring item
N_WEEKS = 13           # weeks shown on the cash flow
SLOTS = 10             # recurring rows shown per week
VAR_ROWS = 6           # weekly variable-expense rows per week

REC_HDR = 4                        # header row on Recurring
REC_FIRST = REC_HDR + 1            # first data row -> item i lives on REC_FIRST+i-1
SCH_HDR = 3
SCH_FIRST = SCH_HDR + 1            # first engine row
SCH_LAST = SCH_FIRST + N_ITEMS * N_OCC - 1

CF_FIRST_BLOCK = 13
BLOCK_H = SLOTS + VAR_ROWS + 5     # header + col hdr + slots + subhdr + vars + totals + spacer


def block_head(w):
    """Row of the 'WEEK n' banner for week w (1-based)."""
    return CF_FIRST_BLOCK + BLOCK_H * (w - 1)


def block_rows(w):
    h = block_head(w)
    return {
        "head": h,
        "cols": h + 1,
        "first_slot": h + 2,
        "last_slot": h + 1 + SLOTS,
        "var_hdr": h + 2 + SLOTS,
        "first_var": h + 3 + SLOTS,
        "last_var": h + 2 + SLOTS + VAR_ROWS,
        "totals": h + 3 + SLOTS + VAR_ROWS,
    }


# ------------------------------------------------------------------- styling
BLUE = "0000FF"        # hardcoded input the user types
BLACK = "000000"       # formula
GREEN = "008000"       # link to another sheet
YELLOW = "FFFF00"      # fill: cells meant to be filled in

F_TITLE = Font(name=FONT, size=16, bold=True, color="1F3864")
F_SUB = Font(name=FONT, size=10, italic=True, color="595959")
F_H1 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_H2 = Font(name=FONT, size=10, bold=True, color="1F3864")
F_BODY = Font(name=FONT, size=10, color=BLACK)
F_INPUT = Font(name=FONT, size=10, color=BLUE)
F_LINK = Font(name=FONT, size=10, color=GREEN)
F_BOLD = Font(name=FONT, size=10, bold=True, color=BLACK)
F_WEEK = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_NOTE = Font(name=FONT, size=9, italic=True, color="595959")

FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_WEEK = PatternFill("solid", fgColor="2E75B6")
FILL_SUB = PatternFill("solid", fgColor="D9E2F3")
FILL_IN = PatternFill("solid", fgColor=YELLOW)
FILL_TOT = PatternFill("solid", fgColor="E2EFDA")
FILL_BAND = PatternFill("solid", fgColor="F2F2F2")
FILL_WARN = PatternFill("solid", fgColor="FCE4D6")
FILL_EMPTY = PatternFill("solid", fgColor="FFF9D6")   # softer "row is free" tint

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPLINE = Border(top=Side(style="medium", color="1F3864"))

MONEY = '$#,##0.00;($#,##0.00);"–"'
MONEY_B = '$#,##0.00;[Red]($#,##0.00);"–"'
DATE_F = 'ddd mm/dd'
DATE_L = 'ddd mmm d, yyyy'

CUR = "'Cash Flow'"


def put(ws, ref, value, font=F_BODY, fmt=None, fill=None, align=None,
        border=None, wrap=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = border
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c


# ------------------------------------------------------- recurring seed data
# Transcribed from the user's three "Recurring" app screenshots (Aug 2026).
# Every Next Due Date is the next occurrence falling on or after START_DATE.
# Anything already received or paid before the sheet opens is rolled forward to
# its following occurrence, so the projection never re-counts money that has
# already moved. The Aug 7 Compunnel paycheck is already in the $71.17 balance.
# type, description, amount, frequency, next due, weekend rule, category
SEED = [
    ("Income",  "Bruno's Restaurant (paycheck)", 90.00,  "Weekly",         dt.date(2026, 8, 12), "None",        "Wages"),
    ("Income",  "Compunnel Software (paycheck)", 790.00, "Weekly",         dt.date(2026, 8, 14), "None",        "Wages"),
    ("Expense", "Apple: Claude",                 21.10,  "Monthly",        dt.date(2026, 9, 4),  "None",        "Subscriptions"),
    ("Expense", "OpenAI",                        20.00,  "Monthly",        dt.date(2026, 9, 5),  "None",        "Subscriptions"),
    ("Expense", "Obsidian",                      10.00,  "Monthly",        dt.date(2026, 9, 5),  "None",        "Subscriptions"),
    ("Expense", "Google One",                     4.99,  "Monthly",        dt.date(2026, 8, 7),  "None",        "Subscriptions"),
    ("Expense", "Upstart (loan)",                500.00, "Monthly",        dt.date(2026, 8, 7),  "None",        "Debt"),
    ("Expense", "Central Maine Power Co.",       163.50, "Monthly",        dt.date(2026, 8, 10), "None",        "Utilities"),
    ("Expense", "Apple: iCloud",                   2.99, "Monthly",        dt.date(2026, 8, 13), "None",        "Subscriptions"),
    ("Expense", "Patrick Rombalski (rent)",      350.00, "Twice a Month",  dt.date(2026, 8, 14), "Move Before", "Housing"),
    ("Expense", "Spectrum",                      171.10, "Monthly",        dt.date(2026, 8, 16), "None",        "Utilities"),
    ("Expense", "Maine Revenue Services",         45.00, "Monthly",        dt.date(2026, 8, 16), "None",        "Taxes"),
    ("Expense", "Metal Copy",                     16.99, "Monthly",        dt.date(2026, 8, 18), "None",        "Subscriptions"),
    ("Expense", "Microsoft",                      13.70, "Monthly",        dt.date(2026, 8, 25), "None",        "Subscriptions"),
    ("Expense", "Central Maine Power: Property",  43.87, "Monthly",        dt.date(2026, 8, 25), "None",        "Utilities"),
    ("Expense", "Maine Revenue Services (2)",    150.00, "Monthly",        dt.date(2026, 8, 30), "None",        "Taxes"),
    ("Expense", "Progressive Insurance",         143.37, "Monthly",        dt.date(2026, 8, 31), "None",        "Insurance"),
]

# Weeks run Friday -> Thursday, so the $790 Friday paycheck lands on day 1 of
# the week rather than in the middle of it. This is the Friday of the week the
# screenshots were taken.
START_DATE = dt.date(2026, 8, 7)
BEGIN_BAL = 71.17                  # the user's actual balance at the start date
CUSHION = 200.00

# No seeded variable expenses: this is a live sheet against a real balance, so
# placeholder spending would distort every ending balance below it.
VAR_SEED = []


wb = Workbook()

# ============================================================== INSTRUCTIONS
ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
for col, wdt in zip("ABCDE", (3, 34, 62, 3, 3)):
    ins.column_dimensions[col].width = wdt

put(ins, "B2", "WEEKLY CASH FLOW SURVIVAL WORKSHEET", F_TITLE)
put(ins, "B3", "Recurring transactions + weekly paychecks, projected 13 weeks out.", F_SUB)

steps = [
    ("HOW TO USE IT", None),
    ("1.  Set your starting point",
     "On the 'Cash Flow' tab, type your Beginning Balance in E5 and your start date in E6 "
     "(use a Friday). Everything below it flows from those two cells. Weeks run "
     "Friday through Thursday, so your Friday paycheck arrives on day one."),
    ("2.  Check your recurring transactions",
     "The 'Recurring' tab is already loaded with the bills and paychecks from your app. "
     "Edit any row, or type a new one into the first empty row - it appears on the cash "
     "flow automatically."),
    ("3.  Record what you actually paid",
     "As each due date arrives or passes, type the real amount into the 'Amount Paid' "
     "column on the Cash Flow tab. The cell turns yellow when something is due and still "
     "blank. Until you fill it in, the sheet projects the scheduled amount."),
    ("4.  Log your week's spending",
     "Each week block has a 'WEEKLY VARIABLE EXPENSES' section. Type a description and an "
     "amount for groceries, gas, and anything else. It is subtracted from that week's "
     "income right away."),
    ("5.  Watch the Summary tab",
     "It shows every week's ending balance, your lowest point, and the first week you run "
     "short - which is the whole point of a survival sheet."),
]
r = 5
for head, body in steps:
    if body is None:
        put(ins, f"B{r}", head, F_H2)
        r += 1
        continue
    put(ins, f"B{r}", head, F_BOLD, align="left")
    put(ins, f"C{r}", body, F_BODY, wrap=True)
    ins.row_dimensions[r].height = 42
    r += 1

r += 1
put(ins, f"B{r}", "COLOUR LEGEND", F_H2)
r += 1
legend = [
    ("Blue text", "You type it. A number or date you enter by hand.", F_INPUT, None),
    ("Yellow fill", "A cell waiting for you to fill it in.", F_BODY, FILL_IN),
    ("Black text", "A formula that does the arithmetic — running balances, week "
                   "totals, the Summary. Leave these alone.", F_BODY, None),
    ("Green text", "A formula that pulls a transaction in from the Recurring tab. "
                   "You CAN type straight over these — see below.", F_LINK, None),
]
for name, desc, fnt, fill in legend:
    put(ins, f"B{r}", name, fnt, fill=fill, border=BOX)
    put(ins, f"C{r}", desc, F_BODY, wrap=True)
    r += 1

r += 1
put(ins, f"B{r}", "WHEN A FORMULA IS IN YOUR WAY", F_H2)
r += 1
overrides = [
    ("Type over it",
     "The green cells in the week blocks — Description, Scheduled Date, Type, Scheduled "
     "Amount — are safe to overwrite. Type a real date or a different amount straight into "
     "the cell. The running balance, the week totals and the Summary all keep working, "
     "because they read the cell rather than the formula behind it."),
    ("What you give up",
     "That one row stops following the Recurring tab. Edit the transaction later and this "
     "row will not move with it. Everything else on the sheet is unaffected."),
    ("To skip a bill this week",
     "Clear the Description cell. The row empties out and drops from the week's totals, "
     "leaving the rest of the week intact."),
    ("To add a one-off",
     "Type a description, date, Income or Expense, and an amount into any blank row in a "
     "week block. Nothing needs to exist on the Recurring tab first."),
    ("To pay part of a bill",
     "Leave the scheduled row alone and put what you actually paid in Amount Paid. That "
     "column is yours already and always wins over the scheduled figure."),
    ("To get the formula back",
     "Undo, or copy the same cell from the row above or below and let Excel adjust it."),
    ("Already paid or already received",
     "Do NOT zero out the row. Go to the Recurring tab and move that transaction's Next "
     "Due Date forward to its following occurrence — that is a blue cell you are meant to "
     "edit, with no formula anywhere near it. The occurrence disappears from the cash flow "
     "and every later one still lands correctly. Your beginning balance already contains "
     "the money, so counting it again would overstate the whole 13 weeks."),
]
for name, desc in overrides:
    put(ins, f"B{r}", name, F_BOLD)
    put(ins, f"C{r}", desc, F_BODY, wrap=True)
    ins.row_dimensions[r].height = 42
    r += 1

r += 1
put(ins, f"B{r}", "FREQUENCIES THE SHEET UNDERSTANDS", F_H2)
r += 1
freqs = [
    ("Weekly", "Every 7 days from the Next Due Date. Your two paychecks use this."),
    ("Every 2 Weeks", "Every 14 days from the Next Due Date."),
    ("Monthly", "Same day each month. A 31st rolls back to the last day of shorter months."),
    ("Twice a Month", "The 15th and the last day of the month."),
    ("One Time", "Happens once, on the Next Due Date, then never again."),
]
for name, desc in freqs:
    put(ins, f"B{r}", name, F_BOLD)
    put(ins, f"C{r}", desc, F_BODY, wrap=True)
    r += 1

r += 1
put(ins, f"B{r}", "ASSUMPTIONS AND SOURCES", F_H2)
r += 1
notes = [
    "Every recurring transaction was transcribed from the three screenshots of your budget "
    "app's Recurring / Upcoming list (August 2026). Nothing was invented.",
    "Amounts are entered as positive numbers. The Type column (Income or Expense) decides "
    "the sign, so a $90 paycheck is Type = Income, amount 90.",
    "The sheet opens on Fri Aug 7, 2026 with $71.17 already in hand. The Aug 7 Compunnel "
    "paycheck had already landed by then, so its Next Due Date is Aug 14 — the first one "
    "still to come. Bruno's, Apple: Claude, OpenAI and Obsidian were rolled forward for the "
    "same reason. Google One ($4.99) and Upstart ($500) are still shown as due on Aug 7; if "
    "those already came out of the account, roll them forward to Sep 7 too.",
    "Progressive Insurance ($143.37) was cut off at the bottom of the third screenshot, so "
    "its due date is set to Aug 31, 2026. Correct it on the Recurring tab if that is wrong.",
    "Patrick Rombalski is 'Twice a Month' with Weekend Rule = Move Before, which is why the "
    "August payment lands on Fri Aug 14 rather than Sat Aug 15, matching your app.",
    "The Beginning Balance of $1,250.00 and the weekly expense examples in Week 1 are "
    "placeholders. Overwrite them with your real numbers.",
    "The Schedule tab is the engine that turns recurring rules into dated occurrences. "
    "It is safe to look at, but do not type in it.",
    "Each week shows up to 10 recurring transactions. If a week ever has more, a warning "
    "appears to the right of that week's banner.",
]
for n in notes:
    put(ins, f"B{r}", "•", F_BODY, align="center")
    put(ins, f"C{r}", n, F_NOTE, wrap=True)
    ins.row_dimensions[r].height = 30
    r += 1
INS_LAST = r - 1


# ================================================================= RECURRING
rec = wb.create_sheet("Recurring")
rec.sheet_view.showGridLines = False
rec_widths = {"A": 6, "B": 30, "C": 11, "D": 12, "E": 16, "F": 15,
              "G": 14, "H": 8, "I": 15, "J": 26}
for col, wdt in rec_widths.items():
    rec.column_dimensions[col].width = wdt

put(rec, "B1", "RECURRING TRANSACTIONS", F_TITLE)
put(rec, "B2", "Edit any row, or add a new one in the first empty row. Blue cells are yours "
               "to type in; the Cash Flow tab updates itself.", F_SUB)

rec_cols = [
    ("A", "ID"), ("B", "Description"), ("C", "Type"), ("D", "Amount"),
    ("E", "Frequency"), ("F", "Next Due Date"), ("G", "Weekend Rule"),
    ("H", "Active"), ("I", "Category"), ("J", "Notes"),
]
for col, label in rec_cols:
    put(rec, f"{col}{REC_HDR}", label, F_H1, fill=FILL_HDR, align="center",
        border=BOX, wrap=True)
rec.row_dimensions[REC_HDR].height = 28

for i in range(1, N_ITEMS + 1):
    row = REC_FIRST + i - 1
    band = FILL_BAND if i % 2 == 0 else None
    put(rec, f"A{row}", i, F_BODY, align="center", border=BOX, fill=band)
    seed = SEED[i - 1] if i <= len(SEED) else None
    typ, desc, amt, freq, due, wknd, cat = seed if seed else ("", "", None, "", None, "", "")
    put(rec, f"B{row}", desc or None, F_INPUT, border=BOX, fill=band)
    put(rec, f"C{row}", typ or None, F_INPUT, align="center", border=BOX, fill=band)
    put(rec, f"D{row}", amt, F_INPUT, fmt=MONEY, border=BOX, fill=band)
    put(rec, f"E{row}", freq or None, F_INPUT, align="center", border=BOX, fill=band)
    put(rec, f"F{row}", due, F_INPUT, fmt=DATE_L, align="center", border=BOX, fill=band)
    put(rec, f"G{row}", wknd or None, F_INPUT, align="center", border=BOX, fill=band)
    put(rec, f"H{row}", "Yes" if seed else None, F_INPUT, align="center", border=BOX, fill=band)
    put(rec, f"I{row}", cat or None, F_INPUT, border=BOX, fill=band)
    put(rec, f"J{row}", None, F_INPUT, border=BOX, fill=band)

REC_LAST = REC_FIRST + N_ITEMS - 1

note_row = REC_LAST + 2
put(rec, f"B{note_row}", "Next Due Date = the next time it happens on or after your cash flow "
                         "start date. Set Active to No to park a transaction without deleting it.",
    F_NOTE, wrap=True)
put(rec, f"B{note_row + 1}",
    "Weekend Rule shifts a due date that lands on a Saturday or Sunday: Move Before = the "
    "Friday, Move After = the Monday, None = leave it alone.", F_NOTE, wrap=True)
put(rec, f"B{note_row + 2}",
    "Source: transcribed from the user's budget-app Recurring screenshots, August 2026.",
    F_NOTE)

dv_type = DataValidation(type="list", formula1='"Income,Expense"', allow_blank=True)
dv_freq = DataValidation(
    type="list",
    formula1='"Weekly,Every 2 Weeks,Monthly,Twice a Month,One Time"', allow_blank=True)
dv_wknd = DataValidation(type="list", formula1='"None,Move Before,Move After"', allow_blank=True)
dv_act = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
for dv, col in ((dv_type, "C"), (dv_freq, "E"), (dv_wknd, "G"), (dv_act, "H")):
    # 'warning', not the default 'stop': the dropdown is a convenience, and it
    # must never be able to block you from typing what you need.
    dv.errorStyle = "warning"
    dv.error = ("The Schedule tab only recognises the listed values, so anything else "
                "will not generate dates. Choose Yes to enter it anyway.")
    dv.errorTitle = "Not one of the listed values"
    rec.add_data_validation(dv)
    dv.add(f"{col}{REC_FIRST}:{col}{REC_LAST}")

# empty rows are the ones to fill in next
rec.conditional_formatting.add(
    f"B{REC_FIRST}:J{REC_LAST}",
    FormulaRule(formula=[f'$B{REC_FIRST}=""'], fill=FILL_EMPTY, stopIfTrue=False))
rec.conditional_formatting.add(
    f"A{REC_FIRST}:J{REC_LAST}",
    FormulaRule(formula=[f'AND($B{REC_FIRST}<>"",$H{REC_FIRST}="No")'],
                font=Font(name=FONT, size=10, italic=True, color="A6A6A6")))
rec.freeze_panes = f"B{REC_FIRST}"


# ================================================================== SCHEDULE
sch = wb.create_sheet("Schedule")
sch.sheet_view.showGridLines = False
for col, wdt in {"A": 7, "B": 7, "C": 30, "D": 11, "E": 12, "F": 14, "G": 14,
                 "H": 9, "I": 8, "J": 16, "K": 12}.items():
    sch.column_dimensions[col].width = wdt

put(sch, "A1", "SCHEDULE ENGINE — DO NOT TYPE IN THIS TAB",
    Font(name=FONT, size=14, bold=True, color="C00000"))
put(sch, "A2", "Expands each recurring rule into dated occurrences and ranks them within "
               "each week. The Cash Flow tab reads the Key column.", F_SUB)

sch_cols = [
    ("A", "Item #"), ("B", "Occ #"), ("C", "Description"), ("D", "Type"),
    ("E", "Amount"), ("F", "Raw Date"), ("G", "Due Date"), ("H", "In Range"),
    ("I", "Week #"), ("J", "Sort Key"), ("K", "Key"),
]
for col, label in sch_cols:
    put(sch, f"{col}{SCH_HDR}", label, F_H1, fill=FILL_HDR, align="center", border=BOX)

for i in range(1, N_ITEMS + 1):
    R = REC_FIRST + i - 1               # this item's row on Recurring
    for k in range(1, N_OCC + 1):
        r = SCH_FIRST + (i - 1) * N_OCC + (k - 1)
        anchor = f"Recurring!$F${R}"
        freq = f"Recurring!$E${R}"
        leg = f"(IF(DAY({anchor})<=20,0,1)+{k - 1})"
        moff = f"INT({leg}/2)"
        semi = (f"IF(MOD({leg},2)=1,EOMONTH({anchor},{moff}),"
                f"DATE(YEAR(EDATE({anchor},{moff})),MONTH(EDATE({anchor},{moff})),15))")

        sch[f"A{r}"] = i
        sch[f"B{r}"] = k
        sch[f"C{r}"] = (f'=IF(OR(Recurring!$B${R}="",Recurring!$H${R}<>"Yes",'
                        f'Recurring!$D${R}=""),"",Recurring!$B${R})')
        sch[f"D{r}"] = f'=IF($C{r}="","",Recurring!$C${R})'
        sch[f"E{r}"] = f'=IF($C{r}="","",Recurring!$D${R})'
        sch[f"F{r}"] = (
            f'=IF($C{r}="","",'
            f'IF({freq}="Weekly",{anchor}+7*{k - 1},'
            f'IF({freq}="Every 2 Weeks",{anchor}+14*{k - 1},'
            f'IF({freq}="Monthly",EDATE({anchor},{k - 1}),'
            f'IF({freq}="Twice a Month",{semi},'
            f'IF({freq}="One Time",IF({k}=1,{anchor},""),""))))))')
        sch[f"G{r}"] = (
            f'=IF($F{r}="","",'
            f'IF(Recurring!$G${R}="Move Before",'
            f'IF(WEEKDAY($F{r},2)=6,$F{r}-1,IF(WEEKDAY($F{r},2)=7,$F{r}-2,$F{r})),'
            f'IF(Recurring!$G${R}="Move After",'
            f'IF(WEEKDAY($F{r},2)=6,$F{r}+2,IF(WEEKDAY($F{r},2)=7,$F{r}+1,$F{r})),'
            f'$F{r})))')
        sch[f"H{r}"] = (f'=IF($G{r}="",0,IF(AND($G{r}>={CUR}!$E$6,'
                        f'$G{r}<{CUR}!$E$6+{N_WEEKS * 7}),1,0))')
        sch[f"I{r}"] = f'=IF($H{r}=1,INT(($G{r}-{CUR}!$E$6)/7)+1,"")'
        sch[f"J{r}"] = f'=IF($H{r}=1,$G{r}*10000+ROW(),"")'
        sch[f"K{r}"] = (f'=IF($H{r}=1,$I{r}*1000+COUNTIFS($I${SCH_FIRST}:$I${SCH_LAST},$I{r},'
                        f'$J${SCH_FIRST}:$J${SCH_LAST},"<"&$J{r})+1,"")')

        for col in "AB":
            sch[f"{col}{r}"].font = F_BODY
            sch[f"{col}{r}"].alignment = Alignment(horizontal="center")
        for col in "CDEFGHIJK":
            sch[f"{col}{r}"].font = F_LINK
        sch[f"E{r}"].number_format = MONEY
        for col in "FG":
            sch[f"{col}{r}"].number_format = DATE_L
            sch[f"{col}{r}"].alignment = Alignment(horizontal="center")
        for col in "HI":
            sch[f"{col}{r}"].alignment = Alignment(horizontal="center")

sch.freeze_panes = f"A{SCH_FIRST}"


# ================================================================= CASH FLOW
cf = wb.create_sheet("Cash Flow", 2)
cf.sheet_view.showGridLines = False
cf_widths = {"A": 3, "B": 32, "C": 13, "D": 11, "E": 14, "F": 14, "G": 14,
             "H": 11, "I": 15, "J": 12, "K": 34}
for col, wdt in cf_widths.items():
    cf.column_dimensions[col].width = wdt
cf.column_dimensions["H"].hidden = True

put(cf, "B1", "WEEKLY CASH FLOW SURVIVAL WORKSHEET", F_TITLE)
put(cf, "B2", "Beginning balance -> every scheduled transaction -> what you actually paid "
              "-> where you stand at the end of each week.", F_SUB)

put(cf, "B4", "STEP 1 — SET YOUR STARTING POINT (type in the yellow cells)", F_H2)
inputs = [
    (5, "Beginning Balance  (cash on hand at the start date)", BEGIN_BAL, MONEY),
    (6, "Cash Flow Start Date  (use a Friday — weeks run Fri to Thu)", START_DATE, DATE_L),
    (7, "Minimum Cash Cushion  (warn me below this)", CUSHION, MONEY),
]
for row, label, val, fmt in inputs:
    put(cf, f"B{row}", label, F_BOLD)
    put(cf, f"E{row}", val, F_INPUT, fmt=fmt, fill=FILL_IN, border=BOX, align="center")
put(cf, "G6", '=IF($E$6="","",IF(WEEKDAY($E$6,2)=5,"",'
              '"! That is a "&TEXT($E$6,"dddd")&". Weeks run Friday to Thursday — '
              'use the Friday on or before the date you want to start."))',
    Font(name=FONT, size=9, bold=True, color="C00000"))
put(cf, "B8", "Today", F_BODY)
put(cf, "E8", "=TODAY()", F_BODY, fmt=DATE_L, align="center", border=BOX)
put(cf, "B9", f"This sheet projects {N_WEEKS} weeks from the start date. "
              f"Week 1 begins on the start date and each week runs Friday "
              f"through Thursday.", F_NOTE)
put(cf, "B10",
    "A yellow 'Amount Paid' cell means that transaction is due or overdue and still "
    "unrecorded.  The green transaction cells are safe to type straight over when a "
    "date or amount is wrong — the balances keep working.  See the Instructions tab.",
    F_NOTE)

paid_ranges, bal_ranges, status_ranges, var_ranges = [], [], [], []

for w in range(1, N_WEEKS + 1):
    b = block_rows(w)
    h, cr = b["head"], b["cols"]
    fs, ls = b["first_slot"], b["last_slot"]
    fv, lv = b["first_var"], b["last_var"]
    tot = b["totals"]
    ws_start = f"$E$6+{7 * (w - 1)}"

    # ---- week banner
    for col in "BCDEFGHIJ":
        cf[f"{col}{h}"].fill = FILL_WEEK
        cf[f"{col}{h}"].font = F_WEEK
    cf[f"B{h}"] = (f'="WEEK {w}     "&TEXT({ws_start},"ddd, mmm d")&"   through   "'
                   f'&TEXT({ws_start}+6,"ddd, mmm d, yyyy")')
    cf[f"B{h}"].alignment = Alignment(horizontal="left", vertical="center")
    cf[f"G{h}"] = "Starting balance:"
    cf[f"G{h}"].alignment = Alignment(horizontal="right", vertical="center")
    cf[f"I{h}"] = "=$E$5" if w == 1 else f"=$I${block_rows(w - 1)['totals']}"
    cf[f"I{h}"].number_format = MONEY_B
    cf[f"I{h}"].alignment = Alignment(horizontal="center", vertical="center")
    cf[f"K{h}"] = (f'=IF(COUNTIF(Schedule!$I${SCH_FIRST}:$I${SCH_LAST},{w})>{SLOTS},'
                   f'"! "&COUNTIF(Schedule!$I${SCH_FIRST}:$I${SCH_LAST},{w})-{SLOTS}'
                   f'&" more transaction(s) fall in this week than there are rows to show them.",'
                   f'"")')
    cf[f"K{h}"].font = Font(name=FONT, size=9, bold=True, color="C00000")
    cf.row_dimensions[h].height = 22

    # ---- column headers
    heads = [("B", "Recurring Transaction"), ("C", "Scheduled Date"), ("D", "Type"),
             ("E", "Scheduled Amount"), ("F", "Amount Paid"), ("G", "Amount Applied"),
             ("H", "Signed"), ("I", "Running Balance"), ("J", "Status")]
    for col, label in heads:
        put(cf, f"{col}{cr}", label, F_H2, fill=FILL_SUB, align="center", border=BOX, wrap=True)
    cf.row_dimensions[cr].height = 26

    # ---- recurring slots
    for n in range(1, SLOTS + 1):
        r = fs + n - 1
        key = w * 1000 + n
        mt = f'MATCH({key},Schedule!$K${SCH_FIRST}:$K${SCH_LAST},0)'
        cf[f"B{r}"] = f'=IFERROR(INDEX(Schedule!$C${SCH_FIRST}:$C${SCH_LAST},{mt}),"")'
        cf[f"C{r}"] = f'=IF($B{r}="","",INDEX(Schedule!$G${SCH_FIRST}:$G${SCH_LAST},{mt}))'
        cf[f"D{r}"] = f'=IF($B{r}="","",INDEX(Schedule!$D${SCH_FIRST}:$D${SCH_LAST},{mt}))'
        cf[f"E{r}"] = f'=IF($B{r}="","",INDEX(Schedule!$E${SCH_FIRST}:$E${SCH_LAST},{mt}))'
        cf[f"F{r}"] = None
        cf[f"G{r}"] = f'=IF($B{r}="","",IF($F{r}<>"",$F{r},$E{r}))'
        cf[f"H{r}"] = f'=IF($B{r}="",0,IF($D{r}="Income",$G{r},-$G{r}))'
        cf[f"I{r}"] = f'=IF($B{r}="","",$I${h}+SUM($H${fs}:$H{r}))'
        cf[f"J{r}"] = (f'=IF($B{r}="","",IF($F{r}<>"","PAID",'
                       f'IF($C{r}<TODAY(),"PAST DUE",'
                       f'IF($C{r}<=TODAY()+7,"DUE","Upcoming"))))')

        for col in "BCDEFGHIJ":
            cf[f"{col}{r}"].border = BOX
            cf[f"{col}{r}"].font = F_LINK if col in "BCD" else F_BODY
        cf[f"F{r}"].font = F_INPUT
        for col in "EFGI":
            cf[f"{col}{r}"].number_format = MONEY
        cf[f"H{r}"].number_format = MONEY
        cf[f"I{r}"].number_format = MONEY_B
        for col in "CDJ":
            cf[f"{col}{r}"].alignment = Alignment(horizontal="center")
        cf[f"C{r}"].number_format = DATE_F

    paid_ranges.append(f"F{fs}:F{ls}")
    status_ranges.append(f"J{fs}:J{ls}")

    # ---- weekly variable expenses
    put(cf, f"B{b['var_hdr']}",
        "WEEKLY VARIABLE EXPENSES — type what you spend this week (groceries, gas, "
        "eating out, anything)", F_H2, fill=FILL_SUB, border=BOX)
    for col in "CDEFGHIJ":
        cf[f"{col}{b['var_hdr']}"].fill = FILL_SUB
        cf[f"{col}{b['var_hdr']}"].border = BOX

    for n in range(VAR_ROWS):
        r = fv + n
        if w == 1 and n < len(VAR_SEED):
            cf[f"B{r}"], cf[f"E{r}"] = VAR_SEED[n]
        cf[f"D{r}"] = f'=IF($B{r}="","","Expense")'
        cf[f"G{r}"] = f'=IF($B{r}="","",N($E{r}))'
        cf[f"H{r}"] = f'=IF($B{r}="",0,-N($E{r}))'
        cf[f"I{r}"] = f'=IF($B{r}="","",$I${h}+SUM($H${fs}:$H{r}))'

        for col in "BCDEFGHIJ":
            cf[f"{col}{r}"].border = BOX
            cf[f"{col}{r}"].font = F_BODY
        for col in "BE":
            cf[f"{col}{r}"].font = F_INPUT
        for col in "EGHI":
            cf[f"{col}{r}"].number_format = MONEY
        cf[f"I{r}"].number_format = MONEY_B
        cf[f"D{r}"].alignment = Alignment(horizontal="center")

    var_ranges.append(f"B{fv}:B{lv}")
    var_ranges.append(f"E{fv}:E{lv}")

    # ---- totals
    put(cf, f"B{tot}", f"WEEK {w} TOTALS", F_BOLD, fill=FILL_TOT, border=BOX)
    put(cf, f"C{tot}", "", F_BODY, fill=FILL_TOT, border=BOX)
    put(cf, f"D{tot}", "", F_BODY, fill=FILL_TOT, border=BOX)
    put(cf, f"E{tot}", f"=SUM($E${fs}:$E${lv})", F_BOLD, fmt=MONEY, fill=FILL_TOT, border=BOX)
    put(cf, f"F{tot}", f"=SUM($F${fs}:$F${ls})", F_BOLD, fmt=MONEY, fill=FILL_TOT, border=BOX)
    put(cf, f"G{tot}", f"=SUM($G${fs}:$G${lv})", F_BOLD, fmt=MONEY, fill=FILL_TOT, border=BOX)
    put(cf, f"H{tot}", f"=SUM($H${fs}:$H${lv})", F_BOLD, fmt=MONEY, fill=FILL_TOT, border=BOX)
    put(cf, f"I{tot}", f"=$I${h}+SUM($H${fs}:$H${lv})", F_BOLD, fmt=MONEY_B,
        fill=FILL_TOT, border=BOX, align="center")
    put(cf, f"J{tot}", f'=IF($I${tot}<0,"SHORTFALL",IF($I${tot}<$E$7,"TIGHT","OK"))',
        F_BOLD, fill=FILL_TOT, border=BOX, align="center")
    put(cf, f"K{tot}",
        f'="Income "&TEXT(SUMIF($D${fs}:$D${lv},"Income",$G${fs}:$G${lv}),"$#,##0.00")'
        f'&"   less expenses "&TEXT(SUMIF($D${fs}:$D${lv},"Expense",$G${fs}:$G${lv}),"$#,##0.00")'
        f'&"   =  net "&TEXT($H${tot},"$#,##0.00")', F_NOTE)
    cf[f"B{tot}"].border = Border(left=THIN, right=THIN, bottom=THIN,
                                  top=Side(style="medium", color="1F3864"))
    bal_ranges.append(f"I{tot}")
    status_ranges.append(f"J{tot}")

# conditional formatting across all week blocks
paid_all = " ".join(paid_ranges)
cf.conditional_formatting.add(
    paid_all,
    FormulaRule(formula=[f'AND($B{CF_FIRST_BLOCK + 2}<>"",$F{CF_FIRST_BLOCK + 2}="",'
                         f'$C{CF_FIRST_BLOCK + 2}<=TODAY())'],
                fill=FILL_IN, border=BOX, stopIfTrue=False))

status_all = " ".join(status_ranges)
cf.conditional_formatting.add(
    status_all,
    FormulaRule(formula=[f'OR($J{CF_FIRST_BLOCK + 2}="PAST DUE",$J{CF_FIRST_BLOCK + 2}="SHORTFALL")'],
                font=Font(name=FONT, size=10, bold=True, color="9C0006"), fill=FILL_WARN))
cf.conditional_formatting.add(
    status_all,
    FormulaRule(formula=[f'OR($J{CF_FIRST_BLOCK + 2}="PAID",$J{CF_FIRST_BLOCK + 2}="OK")'],
                font=Font(name=FONT, size=10, bold=True, color="006100")))
cf.conditional_formatting.add(
    status_all,
    FormulaRule(formula=[f'OR($J{CF_FIRST_BLOCK + 2}="DUE",$J{CF_FIRST_BLOCK + 2}="TIGHT")'],
                font=Font(name=FONT, size=10, bold=True, color="9C6500")))

bal_all = " ".join([f"I{block_rows(w)['first_slot']}:I{block_rows(w)['totals']}"
                    for w in range(1, N_WEEKS + 1)])
cf.conditional_formatting.add(
    bal_all,
    FormulaRule(formula=[f'AND($I{CF_FIRST_BLOCK + 2}<>"",$I{CF_FIRST_BLOCK + 2}<0)'],
                font=Font(name=FONT, size=10, bold=True, color="9C0006")))

cf.freeze_panes = "B12"


# =================================================================== SUMMARY
sm = wb.create_sheet("Summary", 3)
sm.sheet_view.showGridLines = False
for col, wdt in {"A": 3, "B": 9, "C": 20, "D": 15, "E": 14, "F": 16, "G": 16,
                 "H": 13, "I": 16, "J": 12, "K": 11}.items():
    sm.column_dimensions[col].width = wdt

put(sm, "B1", "13-WEEK SURVIVAL SUMMARY", F_TITLE)
put(sm, "B2", "Every figure here is pulled live from the Cash Flow tab.", F_SUB)

SM_HDR = 12
SM_FIRST = SM_HDR + 1
SM_LAST = SM_HDR + N_WEEKS

kpis = [
    ("B4", "Lowest ending balance", f"=MIN($I${SM_FIRST}:$I${SM_LAST})", MONEY_B),
    ("B5", "Week it happens",
     f'="Week "&INDEX($B${SM_FIRST}:$B${SM_LAST},'
     f'MATCH(MIN($I${SM_FIRST}:$I${SM_LAST}),$I${SM_FIRST}:$I${SM_LAST},0))'
     f'&", beginning "&TEXT(INDEX($C${SM_FIRST}:$C${SM_LAST},'
     f'MATCH(MIN($I${SM_FIRST}:$I${SM_LAST}),$I${SM_FIRST}:$I${SM_LAST},0)),"mmm d, yyyy")', None),
    ("B6", "First week you run short",
     f'=IF(MIN($K${SM_FIRST}:$K${SM_LAST})=999,"None — you stay above zero all {N_WEEKS} weeks",'
     f'"Week "&MIN($K${SM_FIRST}:$K${SM_LAST})&"  ("&TEXT(INDEX($C${SM_FIRST}:$C${SM_LAST},'
     f'MATCH(MIN($K${SM_FIRST}:$K${SM_LAST}),$B${SM_FIRST}:$B${SM_LAST},0)),"mmm d, yyyy")&")")', None),
    ("B7", "Total income projected", f"=SUM($E${SM_FIRST}:$E${SM_LAST})", MONEY),
    ("B8", "Total expenses projected",
     f"=SUM($F${SM_FIRST}:$F${SM_LAST})+SUM($G${SM_FIRST}:$G${SM_LAST})", MONEY),
    ("B9", f"Net change over {N_WEEKS} weeks",
     f"=SUM($H${SM_FIRST}:$H${SM_LAST})", MONEY_B),
    ("B10", "Ending balance", f"=$I${SM_LAST}", MONEY_B),
]
for ref, label, formula, fmt in kpis:
    row = int(ref[1:])
    put(sm, f"B{row}", label, F_BOLD)
    sm.merge_cells(f"B{row}:D{row}")
    put(sm, f"E{row}", formula, F_LINK, fmt=fmt, fill=FILL_TOT, border=BOX, align="center")
    if row in (5, 6):
        sm.merge_cells(f"E{row}:I{row}")
        sm[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center")

sm_cols = [("B", "Week"), ("C", "Week Beginning"), ("D", "Starting Balance"),
           ("E", "Income"), ("F", "Recurring Expenses"), ("G", "Weekly Expenses"),
           ("H", "Net Change"), ("I", "Ending Balance"), ("J", "Status"),
           ("K", "shortfall key")]
for col, label in sm_cols:
    put(sm, f"{col}{SM_HDR}", label, F_H1, fill=FILL_HDR, align="center", border=BOX, wrap=True)
sm.row_dimensions[SM_HDR].height = 30
sm.column_dimensions["K"].hidden = True

for w in range(1, N_WEEKS + 1):
    r = SM_HDR + w
    b = block_rows(w)
    fs, ls, fv, lv, tot, h = (b["first_slot"], b["last_slot"], b["first_var"],
                              b["last_var"], b["totals"], b["head"])
    band = FILL_BAND if w % 2 == 0 else None
    put(sm, f"B{r}", w, F_BODY, align="center", border=BOX, fill=band)
    put(sm, f"C{r}", f"={CUR}!$E$6+{7 * (w - 1)}",
        F_LINK, fmt=DATE_L, align="center", border=BOX, fill=band)
    put(sm, f"D{r}", f"={CUR}!$I${h}", F_LINK, fmt=MONEY_B, border=BOX, fill=band)
    put(sm, f"E{r}", f'=SUMIF({CUR}!$D${fs}:$D${ls},"Income",{CUR}!$G${fs}:$G${ls})',
        F_LINK, fmt=MONEY, border=BOX, fill=band)
    put(sm, f"F{r}", f'=SUMIF({CUR}!$D${fs}:$D${ls},"Expense",{CUR}!$G${fs}:$G${ls})',
        F_LINK, fmt=MONEY, border=BOX, fill=band)
    put(sm, f"G{r}", f'=SUMIF({CUR}!$D${fv}:$D${lv},"Expense",{CUR}!$G${fv}:$G${lv})',
        F_LINK, fmt=MONEY, border=BOX, fill=band)
    put(sm, f"H{r}", f"=$E{r}-$F{r}-$G{r}", F_BOLD, fmt=MONEY_B, border=BOX, fill=band)
    put(sm, f"I{r}", f"={CUR}!$I${tot}", F_BOLD, fmt=MONEY_B, border=BOX, fill=band)
    put(sm, f"J{r}", f'=IF($I{r}<0,"SHORTFALL",IF($I{r}<{CUR}!$E$7,"TIGHT","OK"))',
        F_BOLD, align="center", border=BOX, fill=band)
    put(sm, f"K{r}", f"=IF($I{r}<0,$B{r},999)", F_BODY)

TOT_R = SM_LAST + 1
put(sm, f"B{TOT_R}", "TOTAL", F_BOLD, fill=FILL_TOT, border=BOX, align="center")
put(sm, f"C{TOT_R}", "opening →", F_NOTE, fill=FILL_TOT, border=BOX, align="right")
put(sm, f"D{TOT_R}", f"=$D${SM_FIRST}", F_BOLD, fmt=MONEY_B, fill=FILL_TOT, border=BOX)
for col in "EFGH":
    put(sm, f"{col}{TOT_R}", f"=SUM(${col}${SM_FIRST}:${col}${SM_LAST})", F_BOLD,
        fmt=MONEY if col != "H" else MONEY_B, fill=FILL_TOT, border=BOX)
put(sm, f"I{TOT_R}", f"=$I${SM_LAST}", F_BOLD, fmt=MONEY_B, fill=FILL_TOT, border=BOX)
put(sm, f"J{TOT_R}", "", F_BODY, fill=FILL_TOT, border=BOX)

sm.conditional_formatting.add(
    f"B{SM_FIRST}:J{SM_LAST}",
    FormulaRule(formula=[f'$J{SM_FIRST}="SHORTFALL"'], fill=FILL_WARN,
                font=Font(name=FONT, size=10, bold=True, color="9C0006")))
sm.conditional_formatting.add(
    f"J{SM_FIRST}:J{SM_LAST}",
    FormulaRule(formula=[f'$J{SM_FIRST}="TIGHT"'],
                font=Font(name=FONT, size=10, bold=True, color="9C6500")))
sm.conditional_formatting.add(
    f"D{SM_FIRST}:I{TOT_R}",
    FormulaRule(formula=[f'AND($D{SM_FIRST}<>"",$D{SM_FIRST}<0)'],
                font=Font(name=FONT, size=10, bold=True, color="9C0006")))

note = TOT_R + 2
put(sm, f"B{note}", "Recurring Expenses come from the scheduled rows in each week block; "
                    "Weekly Expenses come from what you type into that week's variable-expense "
                    "rows. Both are already netted against income in Net Change.", F_NOTE, wrap=True)
sm.merge_cells(f"B{note}:J{note}")
sm.row_dimensions[note].height = 30
sm.freeze_panes = f"B{SM_FIRST}"


# ------------------------------------------------------------- print set-up
def fit_to_width(ws, landscape=True, area=None, repeat=None):
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    if area:
        ws.print_area = area
    if repeat:
        ws.print_title_rows = repeat


fit_to_width(ins, landscape=False, area=f"A1:C{INS_LAST}")
fit_to_width(rec, area=f"A1:J{note_row + 2}", repeat=f"{REC_HDR}:{REC_HDR}")
fit_to_width(cf, area=f"A1:K{block_rows(N_WEEKS)['totals']}")
fit_to_width(sm, area=f"A1:J{note}")

wb.active = wb.sheetnames.index("Cash Flow")
wb.save(OUT)
print(f"wrote {OUT}")
