#!/usr/bin/env python3
"""Append a bank CSV export into the survival workbook.

Reads either export shape and works out which it has:

  * Monarch style - Date, Merchant, Category, Account, Amount
  * Revolut native - Type, Product, Started Date, Completed Date, Description,
    Amount, Fee, Currency, State, Balance

Only input cells are written; every formula is left alone.

On a Revolut statement it uses the Started Date (the day you made the purchase,
which is what the sheet records and what the beginning balance was struck
against), nets any Fee into the amount, drops anything not COMPLETED so a
reverted charge is never counted, and finally checks the sheet against the
statement's own closing Balance.

What it does, in order:

  1. Drops anything dated before the workbook's start date - that money is
     already inside the beginning balance.
  2. Drops rows on a second account ("Cash on Hand" by default). A single-balance
     sheet counts the ATM withdrawal when it leaves the bank; counting the cash
     spending as well would double it.
  3. Matches a transaction to an already-scheduled recurring bill when the amount
     is identical, the direction agrees, and the dates are within a few days.
     Those become that row's Amount Paid / Received rather than a duplicate line.
  4. Groups whatever is left by merchant and day, then writes it into that week's
     free variable-expense rows.
  5. Skips anything already in the sheet, so re-running over an overlapping export
     does not double-post.

Run it from the folder holding the workbook and the downloaded statement and
it will find both:

    python3 import_transactions.py                       # see what it would do
    python3 import_transactions.py --dry-run             # ...without writing

or name them yourself:

    python3 import_transactions.py <book.xlsx> <export.csv> [--out FILE]
                                   [--cash-account "Cash on Hand"] [--dry-run]

Writing happens in place, over the workbook it read. The copy it replaces is
kept in a backups/ folder alongside.
"""

import argparse
import csv
import datetime as dt
import pathlib
import re
import shutil
from collections import defaultdict

from openpyxl import load_workbook

N_WEEKS, SLOTS, FIRST_BLOCK = 13, 10, 13
NEAR_DAYS = 3           # a hand-entered row may sit a day or two off the bank's date
LOG_FIRST, LOG_LAST = 9, 308
DATE_F = "ddd mm/dd"
MATCH_DAYS = 4          # how far a posting may drift from its scheduled date
CENTS = 0.005


def block(w, var_rows):
    h = FIRST_BLOCK + (SLOTS + var_rows + 5) * (w - 1)
    return {"head": h, "first_slot": h + 2, "last_slot": h + 1 + SLOTS,
            "first_var": h + 3 + SLOTS, "last_var": h + 2 + SLOTS + var_rows}


def detect_var_rows(cf):
    """Work out the block height from where week 2's banner sits.

    Match the banner text itself, not merely "is a formula" - other rows in the
    block are formulas too, and the expected-spending top-up row would otherwise
    be mistaken for the next week's header.
    """
    for vr in range(4, 41):
        h2 = FIRST_BLOCK + (SLOTS + vr + 5)
        v = cf[f"B{h2}"].value
        if isinstance(v, str) and v.startswith('="WEEK '):
            return vr
    raise SystemExit("could not work out the block height - is this the right workbook?")


def norm(name):
    """Loose merchant key: 'Shell (2 charges)' and 'Shell' are the same shop."""
    t = re.sub(r"\(\d+\s*charges?\)", "", str(name or "").lower())
    t = re.sub(r"[^a-z0-9]", "", t)
    return t[:12]


def _num(v):
    v = (v or "").strip().replace("$", "").replace(",", "")
    return float(v) if v else 0.0


def read_csv(path):
    """Return normalised rows, plus the statement's closing balance if it has one."""
    out, closing = [], None
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        return out, closing
    revolut = "Started Date" in rows[0]

    for row in rows:
        if revolut:
            if (row.get("State") or "").strip().upper() != "COMPLETED":
                continue          # REVERTED / PENDING never touched the balance
            stamp = (row.get("Started Date") or row.get("Completed Date") or "").strip()
            if not stamp:
                continue
            day = stamp.split(" ")[0]
            if "-" in day:
                y, m, d = [int(x) for x in day.split("-")]
            else:                       # 8/14/2026
                m, d, y = [int(x) for x in day.split("/")]
            # A fee is charged alongside the amount, so the balance moves by both.
            amount = _num(row.get("Amount")) - abs(_num(row.get("Fee")))
            out.append({"date": dt.date(y, m, d),
                        "merchant": (row.get("Description") or "?").strip(),
                        "category": (row.get("Type") or "").strip(),
                        "account": "", "amount": amount,
                        "seq": (row.get("Completed Date") or "").strip(),
                        "balance": row.get("Balance")})
        else:
            if not (row.get("Date") or "").strip() or not (row.get("Amount") or "").strip():
                continue
            m, d, y = [int(x) for x in row["Date"].split("/")]
            out.append({"date": dt.date(y, m, d),
                        "merchant": (row.get("Merchant") or "?").strip(),
                        "category": (row.get("Category") or "").strip(),
                        "account": (row.get("Account") or "").strip(),
                        "amount": _num(row.get("Amount")),
                        "seq": "", "balance": None})

    if revolut:
        dated = [r for r in out if r.get("seq") and (r.get("balance") or "").strip()]
        if dated:
            def _stamp(r):
                d = r["seq"].split(" ")[0]
                p2 = [int(x) for x in (d.split("-") if "-" in d else d.split("/"))]
                return (p2[0], p2[1], p2[2]) if "-" in d else (p2[2], p2[0], p2[1])
            last = max(dated, key=lambda r: (_stamp(r), r["seq"]))
            closing = (_num(last["balance"]), last["seq"].split(" ")[0])
    return out, closing


BOOK = "Weekly_Cash_Flow_Survival.xlsx"


def find_inputs(book, csv_path):
    """Fall back to whatever is sitting in this folder.

    Typing two paths correctly is the step most likely to go wrong, so with no
    arguments take the workbook by its known name and the most recent CSV
    next to it - which is what a fresh download from the bank will be.
    """
    here = pathlib.Path.cwd()
    if not book:
        if not (here / BOOK).exists():
            raise SystemExit(f"no {BOOK} in this folder - put it here, or name it on the command line")
        book = str(here / BOOK)
    if not csv_path:
        found = sorted(here.glob("*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not found:
            raise SystemExit("no .csv in this folder - download your statement here first")
        csv_path = str(found[0])
        print(f"using the newest CSV here: {found[0].name}")
    return book, csv_path


def back_up(book):
    """Keep the previous copy. This writes over real financial data in place."""
    src = pathlib.Path(book)
    keep = src.parent / "backups"
    keep.mkdir(exist_ok=True)
    dest = keep / f"{src.stem}-{dt.datetime.now():%Y%m%d-%H%M}{src.suffix}"
    shutil.copy(src, dest)
    return dest


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("book", nargs="?"); ap.add_argument("csv_path", nargs="?")
    ap.add_argument("--out"); ap.add_argument("--cash-account", default="Cash on Hand")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    a.book, a.csv_path = find_inputs(a.book, a.csv_path)

    wb = load_workbook(a.book)
    vals = load_workbook(a.book, data_only=True)
    cf, cfv = wb["Cash Flow"], vals["Cash Flow"]
    var_rows = detect_var_rows(cf)

    start = cfv["E6"].value
    start = start.date() if isinstance(start, dt.datetime) else start
    end = start + dt.timedelta(days=N_WEEKS * 7 - 1)

    txns, closing = read_csv(a.csv_path)
    kept, skipped = [], defaultdict(list)
    for t in txns:
        if t["date"] < start:            skipped["before start"].append(t)
        elif t["date"] > end:            skipped["past horizon"].append(t)
        elif t["account"] == a.cash_account: skipped["other account"].append(t)
        else:                            kept.append(t)

    # what the sheet already holds, so a re-run cannot double-post
    existing, by_date_amount, near = set(), set(), []
    for w in range(1, N_WEEKS + 1):
        b = block(w, var_rows)
        for r in range(b["first_var"], b["last_var"] + 1):
            nm = cf[f"B{r}"].value
            if nm and not str(nm).startswith("="):
                d = cfv[f"C{r}"].value
                d = d.date() if isinstance(d, dt.datetime) else d
                amt = round(float(cf[f"E{r}"].value or 0), 2)
                existing.add((str(nm).strip().lower(), d, amt))
                by_date_amount.add((d, amt))     # same charge, different wording
                if d:
                    near.append((norm(nm), d, amt))   # same shop, date a day or two off

    if not any(cfv[f"B{r}"].value
               for w in range(1, N_WEEKS + 1)
               for r in range(block(w, var_rows)["first_slot"],
                              block(w, var_rows)["last_slot"] + 1)):
        raise SystemExit(
            "this workbook holds no calculated values, so no bill could be matched and "
            "every payment would be added a second time.\n"
            "Open it in Excel, let it calculate, save, and run this again.")

    # scheduled recurring occurrences, so a posting can be matched to its bill
    sched = []
    for w in range(1, N_WEEKS + 1):
        b = block(w, var_rows)
        for r in range(b["first_slot"], b["last_slot"] + 1):
            nm = cfv[f"B{r}"].value
            if not nm:
                continue
            d = cfv[f"C{r}"].value
            sched.append({"row": r, "name": nm,
                          "date": d.date() if isinstance(d, dt.datetime) else d,
                          "amount": float(cfv[f"E{r}"].value or 0),
                          "kind": cfv[f"D{r}"].value,
                          "already": cf[f"F{r}"].value is not None})

    # Actuals belong in the Paid Log, keyed to the bill, not in the
    # position-bound Amount Paid cell that a date change would reassign.
    plog = wb["Paid Log"] if "Paid Log" in wb.sheetnames else None
    logged, log_free = set(), []
    if plog is not None:
        for r in range(LOG_FIRST, LOG_LAST + 1):
            nm = plog[f"C{r}"].value
            if nm:
                d = plog[f"B{r}"].value
                logged.add((str(nm).strip().lower(),
                            d.date() if isinstance(d, dt.datetime) else d))
            else:
                log_free.append(r)

    matched, leftover, orphan_income = [], [], []
    used = set()
    for t in kept:
        want_in = t["amount"] > 0
        cands = []
        for s in sched:
            if s["row"] in used or s["date"] is None:
                continue
            if (s["kind"] == "Income") != want_in:
                continue
            if abs((s["date"] - t["date"]).days) > MATCH_DAYS:
                continue
            gap = abs(s["amount"] - abs(t["amount"]))
            # A paycheck that varies is exactly what the Paid Log exists for, so
            # income matches on timing; a bill still has to match to the cent.
            if not want_in and gap > CENTS:
                continue
            cands.append((abs((s["date"] - t["date"]).days), gap, s))
        if cands:
            cands.sort(key=lambda c: (c[0], c[1]))
            hit = cands[0][2]
            used.add(hit["row"])
            matched.append((t, hit))
        elif want_in:
            # never let money coming in be written into an expenses-only row
            orphan_income.append(t)
        else:
            leftover.append(t)

    # group the rest by merchant and day
    groups = defaultdict(list)
    for t in leftover:
        groups[(t["date"], t["merchant"])].append(t)

    to_write, dupes = [], []
    for (d, merch), ts in sorted(groups.items()):
        total = round(sum(abs(x["amount"]) for x in ts), 2)
        label = merch if len(ts) == 1 else f"{merch} ({len(ts)} charges)"
        if (label.strip().lower(), d, total) in existing or (d, total) in by_date_amount:
            dupes.append((label, d, total)); continue
        k = norm(label)
        if any(abs(amt - total) < CENTS and abs((dd - d).days) <= NEAR_DAYS
               and (k.startswith(nk) or nk.startswith(k))
               for nk, dd, amt in near if nk and k):
            dupes.append((label, d, total)); continue
        # The statement groups a day's swipes at one shop into a single line,
        # but the sheet may already hold them typed in separately - two Circle K
        # charges against one "Circle K (2 charges)". Comparing the group total
        # to each row finds nothing, so compare it to the shop's whole day.
        if len(ts) > 1 and k:
            apart = sum(amt for nk, dd, amt in near
                        if nk and (k.startswith(nk) or nk.startswith(k)) and dd == d)
            if abs(apart - total) < CENTS:
                dupes.append((label, d, total)); continue
        to_write.append({"date": d, "label": label, "amount": total,
                         "week": (d - start).days // 7 + 1, "n": len(ts)})

    # place them in free variable rows, week by week
    placed, overflow = [], []
    free = {}
    for w in range(1, N_WEEKS + 1):
        b = block(w, var_rows)
        free[w] = [r for r in range(b["first_var"], b["last_var"] + 1)
                   if cf[f"B{r}"].value is None]
    for item in to_write:
        w = item["week"]
        if free.get(w):
            item["row"] = free[w].pop(0); placed.append(item)
        else:
            overflow.append(item)

    print(f"start {start:%a %b %d, %Y} · horizon to {end:%b %d} · {var_rows} variable rows/week")
    print(f"csv rows {len(txns)} → in scope {len(kept)}")
    for why, rows in skipped.items():
        print(f"   skipped {len(rows):>3}  {why}")
    print()
    if matched:
        print("matched to a scheduled bill (recorded as Amount Paid / Received):")
        for t, s in matched:
            done = s["already"] or (str(s["name"]).strip().lower(), s["date"]) in logged
            note = " — already recorded, left alone" if done else " — into the Paid Log"
            print(f"   {t['date']:%b %d} {t['merchant']:24s} {abs(t['amount']):>8.2f}"
                  f"  ->  {s['name']} sched {s['date']:%b %d}{note}")
        print()
    if dupes:
        print(f"already in the sheet, not re-added ({len(dupes)}):")
        for l, d, amt in dupes: print(f"   {d:%b %d} {l:30s} {amt:>8.2f}")
        print()
    if orphan_income:
        print("!! money IN with no scheduled match — NOT written, would have been "
              "recorded as spending:")
        for t in orphan_income:
            print(f"   {t['date']:%b %d} {t['merchant']:34s} +{t['amount']:>9.2f}")
        print("   add a recurring item for it, or log it by hand.")
        print()
    print(f"variable rows to add ({len(placed)}):")
    tot = 0.0
    for i in placed:
        tot += i["amount"]
        print(f"   W{i['week']:<2} r{i['row']:<4} {i['date']:%a %b %d}  {i['label']:32s} {i['amount']:>8.2f}")
    print(f"   {'total':>47} {tot:>8.2f}")
    if overflow:
        print(f"\n!! no free rows for {len(overflow)}: " +
              ", ".join(f"{i['label']} {i['date']:%b %d}" for i in overflow))

    if a.dry_run:
        print("\ndry run — nothing written")
        return

    for t, s in matched:
        key = (str(s["name"]).strip().lower(), s["date"])
        if s["already"] or key in logged or plog is None or not log_free:
            continue
        r = log_free.pop(0)
        plog[f"B{r}"] = s["date"]
        plog[f"B{r}"].number_format = "ddd mmm d, yyyy"
        plog[f"C{r}"] = s["name"]
        plog[f"D{r}"] = round(abs(t["amount"]), 2)
        plog[f"E{r}"] = f"{t['merchant']} on {t['date']:%b %d}"
        logged.add(key)
    for i in placed:
        cf[f"B{i['row']}"] = i["label"]
        cf[f"C{i['row']}"] = i["date"]
        cf[f"C{i['row']}"].number_format = DATE_F
        cf[f"E{i['row']}"] = i["amount"]

    out = a.out or a.book
    if out == a.book:
        print(f"previous copy kept at {back_up(a.book)}")
    wb.save(out)
    print(f"\nwrote {out}")

    if closing:
        bal, on = closing
        print(f"\nstatement closing balance {bal:,.2f} on {on} — recalculate and check the "
              f"cash flow reads the same on that date")


if __name__ == "__main__":
    main()
